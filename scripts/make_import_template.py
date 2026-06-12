"""產生「專業版」標準匯入範本 import_template.xlsx。
設計目標：客戶只要填資料，不需排版 — 開檔即美觀、欄位清楚、含下拉/格式/範例/說明。
用法: python3 scripts/make_import_template.py -> templates/import_template.xlsx
未來 importer 依「標題」讀取 (非位置)。"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "import_template.xlsx")

# ── 配色：深綠 (吉利) + 金 ──
RED = "1F4D3A"; RED_DK = "153A2B"; GOLD = "B9822F"
TINT = "E9F2EC"; ZEBRA = "F4F9F6"; INK = "1F2A24"; MUTED = "6E827A"; BORDER_C = "D4E5DA"
WHITE = "FFFFFF"
FONT = "Microsoft JhengHei"  # 乾淨中文字型 (Excel 預設可用)

DATA_ROWS = 60  # 預先排版好的空白資料列

thin = Side(style="thin", color=BORDER_C)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet(ws, title, subtitle, cols):
    """cols = list of (header, width, number_format, dropdown_options|None)。"""
    ws.sheet_view.showGridLines = False
    ncol = len(cols)
    last = get_column_letter(ncol)

    # row1 banner
    ws.merge_cells(f"A1:{last}1")
    b = ws["A1"]; b.value = f"  Justinvestment · {title}"
    b.fill = PatternFill("solid", fgColor=RED); b.font = Font(name=FONT, bold=True, color=WHITE, size=15)
    b.alignment = Alignment(vertical="center"); ws.row_dimensions[1].height = 34
    # row2 subtitle
    ws.merge_cells(f"A2:{last}2")
    s = ws["A2"]; s.value = "  " + subtitle
    s.fill = PatternFill("solid", fgColor=TINT); s.font = Font(name=FONT, color=RED_DK, size=10)
    s.alignment = Alignment(vertical="center"); ws.row_dimensions[2].height = 20

    # row3 headers
    hr = 3
    for i, (h, w, _nf, _dd) in enumerate(cols, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = PatternFill("solid", fgColor=RED_DK)
        c.font = Font(name=FONT, bold=True, color=WHITE, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hr].height = 26
    ws.freeze_panes = f"A{hr+1}"

    # pre-formatted empty data rows (zebra + border + number format)
    for r in range(hr + 1, hr + 1 + DATA_ROWS):
        fill = PatternFill("solid", fgColor=WHITE if (r % 2 == 0) else ZEBRA)
        for i, (_h, _w, nf, _dd) in enumerate(cols, 1):
            c = ws.cell(row=r, column=i)
            c.fill = fill; c.border = BORDER
            c.font = Font(name=FONT, size=11, color=INK)
            c.alignment = Alignment(vertical="center")
            if nf:
                c.number_format = nf
        ws.row_dimensions[r].height = 20

    # dropdowns + date validation + numeric hints
    for i, (h, _w, nf, dd) in enumerate(cols, 1):
        col = get_column_letter(i)
        rng = f"{col}{hr+1}:{col}{hr+DATA_ROWS}"
        if dd:
            dv = DataValidation(type="list", formula1='"' + ",".join(dd) + '"', allow_blank=True)
            dv.showInputMessage = True
            dv.promptTitle = h.replace("\n", " ").strip()
            dv.prompt = "請從下拉選單選擇"
            ws.add_data_validation(dv); dv.add(rng)
        elif nf == DATE:
            dv = DataValidation(type="date", operator="between",
                                formula1="2000-01-01", formula2="2100-12-31", allow_blank=True)
            dv.showInputMessage = True
            dv.promptTitle = h.replace("\n", " ").strip()
            dv.prompt = "請輸入日期，格式 YYYY-MM-DD（例 2026-06-18）"
            dv.showErrorMessage = True
            dv.errorTitle = "日期格式不正確"
            dv.error = "請用 YYYY-MM-DD（例 2026-06-18）"
            ws.add_data_validation(dv); dv.add(rng)
        elif nf == PCT:
            dv = DataValidation(type="decimal", operator="between",
                                formula1="0", formula2="100000", allow_blank=True)
            dv.showInputMessage = True
            dv.promptTitle = h.replace("\n", " ").strip()
            dv.prompt = "填數字即可（例 KO=100、票息=8 代表 8%）"
            ws.add_data_validation(dv); dv.add(rng)
    return hr


def put_sample(ws, hr, values):
    r = hr + 1
    for i, v in enumerate(values, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.fill = PatternFill("solid", fgColor="F3E6CC")  # soft gold
        c.font = Font(name=FONT, italic=True, color=MUTED, size=11)
        c.border = BORDER


CATEGORIES = ["SN", "台股", "期貨", "美股", "港股", "國內外基金", "ELN", "PGN", "儲蓄險", "旅平險", "車險", "意外險"]
CURRENCIES = ["USD", "TWD", "HKD", "EUR", "JPY", "CNY"]
FREQ = ["月配", "季配", "半年配", "年配", "到期一次"]
STATUS = ["進行中", "已出場", "暫停"]
NUM = "#,##0"; NUM2 = "#,##0.00"; DATE = "yyyy-mm-dd"; PCT = "0"

wb = Workbook()

# ── 說明 (封面) ──
ws0 = wb.active; ws0.title = "說明"; ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 4; ws0.column_dimensions["B"].width = 96
ws0.merge_cells("A1:B2")
t = ws0["A1"]; t.value = "  Justinvestment · 資料匯入範本"
t.fill = PatternFill("solid", fgColor=RED); t.font = Font(name=FONT, bold=True, color=WHITE, size=18)
t.alignment = Alignment(vertical="center"); ws0.row_dimensions[1].height = 30; ws0.row_dimensions[2].height = 16

blocks = [
    ("", ""),
    ("填寫方式", "h"),
    ("只需在各分頁輸入資料即可，版面已排好、無需美化。灰金色斜體列為「範例」，匯入前請刪除。", ""),
    ("", ""),
    ("分頁說明", "h"),
    ("客戶 — 投資人 (你服務的客戶)。必填：姓名。", ""),
    ("商品 — 結構型商品 (SN/ELN/PGN…)。必填：商品代號。SN 類才需填 KO/KI/票息/標的。", ""),
    ("投資 — 將投資人連到商品 (持倉)。必填：客戶姓名、商品代號、投資金額。", ""),
    ("", ""),
    ("格式規定", "h"),
    ("日期：YYYY-MM-DD（例 2026-06-18）", ""),
    ("百分比欄（KO/KI/履約/票息）：填數字即可，例 KO=100、KI=60、票息=8（代表 8% 年化）", ""),
    ("金額：純數字，勿加逗號或貨幣符號", ""),
    ("配息頻率：月配 / 季配 / 半年配 / 年配 / 到期一次（可用下拉選）", ""),
    ("「投資」分頁的客戶姓名，需與「客戶」分頁完全一致才能正確連結", ""),
]
r = 4
for txt, kind in blocks:
    cell = ws0.cell(row=r, column=2, value=txt)
    if kind == "h":
        cell.font = Font(name=FONT, bold=True, size=12, color=RED_DK)
    else:
        cell.font = Font(name=FONT, size=11, color=INK)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws0.row_dimensions[r].height = 22 if txt else 8
    r += 1

# ── 客戶 ──
ws1 = wb.create_sheet("客戶")
hr1 = style_sheet(ws1, "客戶 (投資人)", "你服務的投資人名單 · 必填：姓名", [
    ("姓名 *", 18, None, None), ("美元額度", 14, NUM, None), ("幣別", 9, None, CURRENCIES),
    ("統一帳號", 16, None, None), ("備註", 32, None, None),
])
put_sample(ws1, hr1, ["蔣太太", 370000, "USD", "", "← 範例，請刪除"])

# ── 商品 ──
ws2 = wb.create_sheet("商品")
hr2 = style_sheet(ws2, "商品 (結構型商品)", "SN/ELN/PGN 等 · 必填：商品代號 · % 欄填數字 (KO=100…)", [
    ("商品代號 *", 16, None, None), ("類別", 11, None, CATEGORIES),
    ("標的1", 9, None, None), ("期初價1", 10, NUM2, None),
    ("標的2", 9, None, None), ("期初價2", 10, NUM2, None),
    ("標的3", 9, None, None), ("期初價3", 10, NUM2, None),
    ("KO障壁\n(%)", 9, PCT, None), ("KI障壁\n(%)", 9, PCT, None), ("履約價\n(%)", 9, PCT, None),
    ("票息\n(%年化)", 10, PCT, None), ("配息頻率", 11, None, FREQ),
    ("成交日", 13, DATE, None), ("比價日", 13, DATE, None), ("到期日", 13, DATE, None),
    ("狀態", 10, None, STATUS),
])
put_sample(ws2, hr2, ["EQDS0702653", "SN", "TSLA", 250.5, "TSM", 180.2, "ANET", 95.0,
                      100, 60, 100, 8, "月配", "2026-05-11", "2026-06-18", "", "進行中"])

# ── 投資 ──
ws3 = wb.create_sheet("投資")
hr3 = style_sheet(ws3, "投資 (持倉)", "把投資人連到商品 · 三欄皆必填 · 客戶姓名須與「客戶」分頁一致", [
    ("客戶姓名 *", 18, None, None), ("商品代號 *", 16, None, None),
    ("投資金額 *", 14, NUM, None), ("幣別", 9, None, CURRENCIES),
])
put_sample(ws3, hr3, ["蔣太太", "EQDS0702653", 370000, "USD"])

# ── 連動下拉：投資分頁從「客戶/商品」分頁選 (不必重打)；標的給常用清單 ──
from openpyxl.workbook.defined_name import DefinedName
dnI = DefinedName(name="InvestorList", attr_text="'客戶'!$A$4:$A$63")
dnP = DefinedName(name="ProductList", attr_text="'商品'!$A$4:$A$63")
try:
    wb.defined_names["InvestorList"] = dnI
    wb.defined_names["ProductList"] = dnP
except TypeError:
    wb.defined_names.add(dnI); wb.defined_names.add(dnP)

TICKERS = ["NVDA", "TSLA", "TSM", "ANET", "AMD", "INTC", "AVGO", "MSFT", "ORCL", "MU",
           "GOOG", "GOOGL", "CRCL", "LITE", "COHR", "QQQ", "SPY", "SMH", "SOXX", "IBB",
           "AAPL", "AMZN", "META", "NFLX"]


def add_dv(ws, col, formula1, strict, prompt):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.showErrorMessage = strict
    dv.showInputMessage = True
    dv.prompt = prompt
    ws.add_data_validation(dv)
    dv.add(f"{col}{hr2 + 1 if ws is ws2 else hr3 + 1}:{col}{(hr2 if ws is ws2 else hr3) + DATA_ROWS}")


tk = '"' + ",".join(TICKERS) + '"'
for col in ("C", "E", "G"):  # 標的1/2/3 — เลือกหุ้นที่ใช้บ่อย หรือพิมพ์เอง
    add_dv(ws2, col, tk, False, "可選常用標的，或自行輸入其他代號")
add_dv(ws3, "A", "InvestorList", True, "從「客戶」分頁選擇，不必重打")
add_dv(ws3, "B", "ProductList", True, "從「商品」分頁選擇，不必重打")

wb.save(OUT)
print("wrote", OUT)
