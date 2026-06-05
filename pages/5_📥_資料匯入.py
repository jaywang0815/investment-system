"""
Excel 資料匯入頁面
- อัปโหลดได้หลายไฟล์พร้อมกัน
- รองรับทุกเดือน (อัตโนมัติ)
- Preview ก่อน import
"""
import streamlit as st
import pandas as pd
from io import BytesIO
from utils.excel_parser import parse_excel_file, get_summary, is_sn_sheet, is_customer_sheet

st.set_page_config(page_title="資料匯入", page_icon="📥", layout="wide")

def _is_logged_in():
    if st.session_state.get("authenticated"):
        return True
    try:
        return st.user.is_logged_in
    except Exception:
        return False

if not _is_logged_in():
    st.error("請先登入")
    st.page_link("app.py", label="回到登入頁面", icon="🔑")
    st.stop()

st.title("📥 Excel 資料匯入")
st.caption("支援多個檔案同時匯入，自動識別月份")

try:
    from utils.database import get_supabase
    sb = get_supabase()
    DB_READY = True
except Exception:
    DB_READY = False
    st.warning("⚠️ 資料庫未連線 — 仍可預覽資料，但無法匯入。請先完成「系統設定」。")

def _snapshot_month(sb, month_label: str) -> dict:
    """บันทึก snapshot ของข้อมูลเดือนก่อน import เพื่อใช้ UNDO"""
    sns = sb.table("structured_notes").select("*").eq("month_label", month_label).execute().data or []
    snapshot = {"month_label": month_label, "sns": []}
    for sn in sns:
        invs = sb.table("investments").select("*").eq("sn_id", sn["id"]).execute().data or []
        snapshot["sns"].append({"sn": sn, "investments": invs})
    return snapshot


def _restore_snapshot(sb, snapshot: dict):
    """คืนค่าข้อมูลจาก snapshot (UNDO)"""
    month = snapshot["month_label"]
    status = st.empty()

    # ลบข้อมูลเดือนนั้นทั้งหมดก่อน
    status.text("復原中：刪除現有資料...")
    existing = sb.table("structured_notes").select("id").eq("month_label", month).execute().data or []
    for sn in existing:
        sb.table("investments").delete().eq("sn_id", sn["id"]).execute()
        sb.table("structured_notes").delete().eq("id", sn["id"]).execute()

    # คืนค่าจาก snapshot
    for item in snapshot["sns"]:
        sn_data = {k: v for k, v in item["sn"].items() if k != "id"}
        status.text(f"復原：{item['sn'].get('product_code')}")
        resp = sb.table("structured_notes").insert(sn_data).execute()
        if resp.data:
            new_sn_id = resp.data[0]["id"]
            for inv in item["investments"]:
                inv_data = {k: v for k, v in inv.items() if k not in ("id", "sn_id", "created_at")}
                inv_data["sn_id"] = new_sn_id
                sb.table("investments").insert(inv_data).execute()

    status.empty()
    st.success("✅ 已復原至匯入前的狀態")
    st.session_state.pop("undo_snapshot", None)
    st.rerun()


def _do_import(parsed_list: list, import_customers: bool, import_sns: bool, skip_duplicates: bool, force_month: str = None):
    """執行匯入到 Supabase — รองรับ upsert และ UNDO"""
    from utils.database import get_supabase
    sb = get_supabase()

    upsert_mode = bool(force_month)  # เปิด upsert เมื่อระบุเดือนชัดเจน
    total_customers = 0
    total_sns = 0
    total_updated = 0
    total_investments = 0
    errors = []

    existing_customers = {}
    try:
        resp = sb.table("customers").select("id,name").execute()
        existing_customers = {c["name"]: c["id"] for c in (resp.data or [])}
    except Exception:
        pass

    # บันทึก snapshot ก่อน import (สำหรับ UNDO)
    if upsert_mode and import_sns:
        with st.spinner("備份資料中 (準備復原功能)..."):
            st.session_state["undo_snapshot"] = _snapshot_month(sb, force_month)

    progress = st.progress(0)
    status = st.empty()

    for parsed in parsed_list:
        if import_customers:
            for cust in parsed.get("customers", []):
                name = cust["name"]
                if name in existing_customers:
                    status.text(f"跳過重複客戶: {name}")
                    continue
                try:
                    resp = sb.table("customers").insert(cust).execute()
                    if resp.data:
                        existing_customers[name] = resp.data[0]["id"]
                        total_customers += 1
                        status.text(f"✅ 新增客戶: {name}")
                except Exception as e:
                    errors.append(f"客戶 {name}: {e}")

        if import_sns:
            all_sns = []
            for month_label, month_sns in parsed.get("sn_by_month", {}).items():
                for sn in month_sns:
                    if force_month:
                        sn["month_label"] = force_month
                    all_sns.append(sn)

            for sn_idx, sn in enumerate(all_sns):
                progress.progress((sn_idx + 1) / max(len(all_sns), 1))
                investments = sn.pop("investments", [])
                code = sn["product_code"]

                sn_id = None
                # เช็กว่ามีอยู่แล้วไหม
                try:
                    resp = sb.table("structured_notes").select("id").eq("product_code", code).execute()
                    existing_sn = resp.data[0] if resp.data else None
                except Exception:
                    existing_sn = None

                if existing_sn:
                    if upsert_mode:
                        # อัพเดทข้อมูลเดิม
                        sn_id = existing_sn["id"]
                        sb.table("structured_notes").update(sn).eq("id", sn_id).execute()
                        # ลบ investments เก่าแล้วใส่ใหม่
                        sb.table("investments").delete().eq("sn_id", sn_id).execute()
                        total_updated += 1
                        status.text(f"🔄 更新 SN: {code}")
                    else:
                        status.text(f"跳過重複商品: {code}")
                        continue
                else:
                    try:
                        resp = sb.table("structured_notes").insert(sn).execute()
                        if resp.data:
                            sn_id = resp.data[0]["id"]
                            total_sns += 1
                            status.text(f"✅ 新增 SN: {code}")
                    except Exception as e:
                        errors.append(f"SN {code}: {e}")
                        continue

                if sn_id:
                    for inv in investments:
                        cname = inv["customer_name"]
                        amount = inv["amount_usd"]
                        customer_id = existing_customers.get(cname)
                        if not customer_id:
                            name_clean = cname.replace("*", "").replace("＊", "").strip()
                            for k, v in existing_customers.items():
                                if name_clean in k.replace("*","").replace("＊","").strip() or \
                                   k.replace("*","").replace("＊","").strip() in name_clean:
                                    customer_id = v
                                    break
                        if not customer_id:
                            try:
                                resp = sb.table("customers").insert({"name": cname}).execute()
                                if resp.data:
                                    customer_id = resp.data[0]["id"]
                                    existing_customers[cname] = customer_id
                            except Exception:
                                pass
                        if customer_id:
                            try:
                                sb.table("investments").insert({
                                    "customer_id": customer_id,
                                    "sn_id": sn_id,
                                    "amount_usd": amount
                                }).execute()
                                total_investments += 1
                            except Exception as e:
                                if "duplicate" not in str(e).lower():
                                    errors.append(f"投資記錄 {cname}→{code}: {e}")

    progress.progress(1.0)
    status.empty()

    st.markdown("---")
    st.markdown("### 📊 匯入結果")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("新增客戶", f"{total_customers} 人")
    with c2:
        st.metric("新增 SN", f"{total_sns} 筆")
    with c3:
        st.metric("更新 SN", f"{total_updated} 筆")
    with c4:
        st.metric("錯誤", f"{len(errors)} 個")

    if errors:
        with st.expander(f"⚠️ {len(errors)} 個錯誤"):
            for e in errors:
                st.text(f"• {e}")
    else:
        st.success("✅ 匯入完成，無錯誤！")
        if upsert_mode:
            st.info("💡 如需復原，請點擊下方「↩️ 復原上次匯入」按鈕")
        else:
            st.balloons()

    st.session_state.pop("parsed_data", None)
    st.session_state.pop("parsed_data_list", None)
    st.session_state.pop("use_existing", None)


# ── ปุ่ม UNDO (แสดงเมื่อมี snapshot) ─────────────────────────
if st.session_state.get("undo_snapshot") and DB_READY:
    month = st.session_state["undo_snapshot"]["month_label"]
    st.warning(f"⚠️ มี snapshot ของ **{month}** ก่อน import ล่าสุด")
    col_u1, col_u2 = st.columns([1, 4])
    with col_u1:
        if st.button("↩️ 復原上次匯入", type="primary", use_container_width=True):
            from utils.database import get_supabase as _gsb
            _restore_snapshot(_gsb(), st.session_state["undo_snapshot"])
    with col_u2:
        if st.button("🗑️ 放棄復原 (確認保留新資料)", use_container_width=True):
            st.session_state.pop("undo_snapshot", None)
            st.rerun()
    st.markdown("---")

tab1, tab2 = st.tabs(["📁 上傳並匯入", "📋 月份管理"])

# ═══════════════════════════════════════════════════════════════
# Tab 1: 上傳並匯入
# ═══════════════════════════════════════════════════════════════
with tab1:

    # ── 說明區塊 ──────────────────────────────────────────────
    with st.expander("📖 使用說明 (點擊展開)", expanded=False):
        st.markdown("""
        **支援的 Excel 格式:**
        - 單一檔案含多個月份 Sheet (例如: 開戶明細 + ＳＮ5月 + ＳＮ6月)
        - 多個獨立月份檔案 (例如: 5月.xlsx + 6月.xlsx)
        - 每次可同時上傳多個檔案

        **Sheet 命名規則:**
        - 客戶資料: 含「開戶」或「客戶」字樣
        - SN 商品: 含「ＳＮ」或「SN」+ 月份，如「ＳＮ5月」「SN6月」

        **匯入說明:**
        - 系統會自動偵測月份，不需要手動填寫
        - 重複的客戶或商品代號不會重複匯入
        - 可先預覽資料再決定是否匯入
        """)

    # ── 檔案上傳區 ────────────────────────────────────────────
    uploaded_files = st.file_uploader(
        "📂 拖曳 Excel 檔案到這裡，或點擊選擇",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="可同時選擇多個檔案"
    )

    if not uploaded_files:
        st.info("👆 請上傳 Excel 檔案開始匯入")
        st.markdown("---")
        st.markdown("**快速上傳最近的檔案:**")
        if st.button("📂 使用現有的 開戶明細 金額.xlsx"):
            st.session_state["use_existing"] = True

        if st.session_state.get("use_existing"):
            import os
            existing_path = "/Users/jay/Desktop/開戶明細 金額.xlsx"
            if os.path.exists(existing_path):
                with open(existing_path, "rb") as f:
                    data = f.read()
                # Parse it
                parsed = parse_excel_file(BytesIO(data))
                st.session_state["parsed_data"] = parsed
                st.session_state["file_names"] = ["開戶明細 金額.xlsx"]
                st.rerun()

    # 處理上傳的檔案
    if uploaded_files:
        all_parsed = []
        with st.spinner("📂 讀取檔案中，請稍候..."):
            for file in uploaded_files:
                try:
                    file_bytes = BytesIO(file.read())
                    parsed = parse_excel_file(file_bytes)
                    parsed["_filename"] = file.name
                    all_parsed.append(parsed)
                except Exception as e:
                    st.error(f"❌ 無法讀取 {file.name}: {e}")

        st.session_state["parsed_data_list"] = all_parsed

    # 使用已存在的快取資料
    if st.session_state.get("parsed_data"):
        all_parsed = [st.session_state["parsed_data"]]
        all_parsed[0]["_filename"] = "開戶明細 金額.xlsx"

    parsed_list = st.session_state.get("parsed_data_list") or (
        [st.session_state["parsed_data"]] if st.session_state.get("parsed_data") else []
    )

    if parsed_list:
        st.markdown("---")

        for parsed in parsed_list:
            filename = parsed.get("_filename", "未知檔案")
            summary = get_summary(parsed)

            st.markdown(f"### 📄 {filename}")

            # Sheet 偵測結果
            sheets = parsed.get("sheets_found", {})
            col_s1, col_s2, col_s3 = st.columns(3)
            with col_s1:
                st.markdown(f"**客戶 Sheet:** {', '.join(sheets.get('customer_sheets', [])) or '未偵測到'}")
            with col_s2:
                st.markdown(f"**SN Sheet:** {', '.join(sheets.get('sn_sheets', [])) or '未偵測到'}")
            with col_s3:
                st.markdown(f"**其他 Sheet:** {', '.join(sheets.get('other_sheets', [])) or '無'}")

            # 摘要卡片
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.metric("客戶數", summary["customers"])
            with c2:
                st.metric("月份數", len(summary["months"]))
            with c3:
                st.metric("SN 商品數", summary["total_sns"])
            with c4:
                st.metric("投資記錄", summary["total_investments"])

            if summary["months"]:
                st.markdown(f"**偵測到的月份:** {'、'.join(sorted(summary['months']))}")

            # ── 資料預覽 ──────────────────────────────────────
            with st.expander("👀 預覽客戶資料"):
                customers = parsed.get("customers", [])
                if customers:
                    df = pd.DataFrame(customers)
                    bool_cols = ["unified_account", "pi_signed", "ordered"]
                    for col in bool_cols:
                        if col in df.columns:
                            df[col] = df[col].map({True: "✅", False: "❌"})
                    df = df.rename(columns={
                        "name": "姓名", "unified_account": "統一開戶",
                        "pi_signed": "PI見簽", "ordered": "已下單",
                        "usd_amount": "USD額度", "ctbc_position": "中信部位"
                    })
                    st.dataframe(df, use_container_width=True, hide_index=True)
                else:
                    st.info("無客戶資料")

            for month, sns in sorted(parsed.get("sn_by_month", {}).items()):
                with st.expander(f"👀 預覽 {month} SN 商品 ({len(sns)} 筆)"):
                    preview_rows = []
                    for sn in sns:
                        tickers = " / ".join([sn.get(f"underlying_{i}") for i in range(1, 4)
                                               if isinstance(sn.get(f"underlying_{i}"), str)])
                        preview_rows.append({
                            "代號": sn["product_code"],
                            "標的": tickers,
                            "執行價%": f"{sn['strike_pct']*100:.1f}%" if sn.get("strike_pct") else "—",
                            "配息%": f"{sn['coupon_pct']*100:.2f}%" if sn.get("coupon_pct") else "—",
                            "比價日": str(sn.get("observation_date", ""))[:10],
                            "投資客戶": ", ".join([i["customer_name"] for i in sn.get("investments", [])]),
                        })
                    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

            st.markdown("---")

        # ── 匯入按鈕 ────────────────────────────────────────────
        st.markdown("### ▶️ 開始匯入")

        import_opts = st.columns(3)
        with import_opts[0]:
            import_customers = st.checkbox("匯入客戶資料", value=True)
        with import_opts[1]:
            import_sns = st.checkbox("匯入 SN 商品", value=True)
        with import_opts[2]:
            skip_duplicates = st.checkbox("跳過重複資料", value=True,
                                          help="相同代號或姓名的資料不會重複匯入")

        # ── 指定月份 ─────────────────────────────────────────────
        st.markdown("**📅 指定匯入月份 (選填)**")
        month_options = ["自動偵測 (依 Sheet 名稱)"] + [f"{i}月" for i in range(1, 13)]
        override_month = st.selectbox(
            "若 Excel 沒有月份名稱，或想強制指定月份，請在此選擇",
            month_options,
            help="選「自動偵測」表示依照 Sheet 名稱判斷月份"
        )
        force_month = None if override_month == "自動偵測 (依 Sheet 名稱)" else override_month

        if not DB_READY:
            st.error("❌ 請先完成資料庫設定才能匯入")
        else:
            if st.button("🚀 確認匯入", type="primary", use_container_width=True):
                _do_import(parsed_list, import_customers, import_sns, skip_duplicates, force_month)


# ═══════════════════════════════════════════════════════════════
# Tab 2: 月份管理
# ═══════════════════════════════════════════════════════════════
with tab2:
    st.subheader("📅 月份資料管理")

    st.markdown("""
    **如何新增每月資料:**

    每個月收到銀行的新 SN 商品後，有兩種方式更新:

    **方式 A: 在現有 Excel 新增 Sheet**
    1. 開啟「開戶明細 金額.xlsx」
    2. 在底部新增一個 Sheet，命名為「ＳＮ6月」(或其他月份)
    3. 按照相同格式填入新商品資料
    4. 回到「上傳並匯入」頁面，上傳整個檔案

    **方式 B: 建立新的月份 Excel 檔**
    1. 建立新的 Excel 檔案，只包含新月份的 SN 資料
    2. Sheet 命名為「ＳＮ6月」等
    3. 上傳此新檔案即可，不會影響已匯入的資料

    **方式 C: 直接在系統新增 (推薦)**
    → 前往「SN商品管理」頁面，點選「新增商品」直接填寫
    """)

    st.markdown("---")

    if DB_READY:
        st.subheader("已匯入的月份資料")
        try:
            from utils.database import get_all_sns
            sns_df = get_all_sns()
            if not sns_df.empty and "month_label" in sns_df.columns:
                month_counts = sns_df.groupby("month_label").size().reset_index(name="商品數量")
                month_counts.columns = ["月份", "商品數量"]
                st.dataframe(month_counts, use_container_width=True, hide_index=True)
            else:
                st.info("尚無資料")
        except Exception:
            st.info("請先完成資料庫設定")

    st.markdown("---")

    # 下載 Excel 範本
    st.subheader("📥 下載 Excel 範本")
    st.markdown("如果要建立新月份的 Excel 檔，可以下載範本參考格式:")

    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb_template = openpyxl.Workbook()

        # 客戶 Sheet
        ws_c = wb_template.active
        ws_c.title = "開戶明細"
        headers_c = ["戶名", "統一開戶", "ＰＩ見簽", "已下單", "ＵＳＤ", "中信部位", "FUND"]
        for j, h in enumerate(headers_c, 1):
            cell = ws_c.cell(row=1, column=j, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
            cell.font = Font(bold=True, color="FFFFFF")
        ws_c.append(["範例客戶", "Ｖ", "Ｖ", "Ｖ", 500000, None, None])

        # SN Sheet 範本
        ws_sn = wb_template.create_sheet("ＳＮ6月")
        headers_sn = ["日期", "代號", "標的1", "標的2", "標的3", "標的4", "標的5",
                       "執行價", "配息", "比價", "KO提前", "KI下限", "出場", "暫結", "CHU", "下單金額"]
        for j, h in enumerate(headers_sn, 1):
            cell = ws_sn.cell(row=1, column=j, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
        # 範例資料行
        from datetime import date as dt
        ws_sn.append([dt.today(), "EQDS0106001", "NVDA", "TSLA", "AAPL", None, None,
                       0.80, 0.15, dt.today(), 1.0, 0.5, None, None, None, 500000])
        ws_sn.append([None, "期初價格", 220.0, 350.0, 200.0, None, None,
                       None, None, None, None, None, None, None, None, None])
        ws_sn.append(["客戶姓名", 200000, None, None, None, None, None,
                       None, None, None, None, None, None, None, None, None])

        buf = io.BytesIO()
        wb_template.save(buf)
        buf.seek(0)

        st.download_button(
            label="⬇️ 下載 Excel 範本 (ＳＮ6月格式)",
            data=buf.getvalue(),
            file_name="SN月份範本.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.warning(f"無法產生範本: {e}")
