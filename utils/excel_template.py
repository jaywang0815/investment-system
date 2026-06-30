"""統一的匯入/匯出 Excel 範本產生器 (深綠吉利 + 金)。
build_workbook(data=None, with_sample=False) →
  - data=None: 空白範本 (含範例列)
  - data={customers,products,investments}: 已填資料 + 下方仍有空白列可續填 + 下拉/驗證
匯入器依「標題」讀取，因此匯出檔可再匯入 (round-trip)。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

GREEN = "1F4D3A"; GREEN_DK = "153A2B"; GOLD = "B9822F"
TINT = "E9F2EC"; ZEBRA = "F4F9F6"; INK = "1F2A24"; MUTED = "6E827A"; BORDER_C = "D4E5DA"; WHITE = "FFFFFF"
FONT = "Microsoft JhengHei"
BLANK_ROWS = 80
NUM = "#,##0"; NUM2 = "#,##0.00"; DATE = "yyyy-mm-dd"; PCT = "0"

CATEGORIES = ["SN", "台股", "期貨", "美股", "港股", "國內外基金", "ELN", "PGN", "儲蓄險", "旅平險", "車險", "意外險"]
CURRENCIES = ["USD", "TWD", "HKD", "EUR", "JPY", "CNY"]
FREQ = ["月配", "季配", "半年配", "年配", "到期一次"]
STATUS = ["進行中", "已出場", "暫停"]
TICKERS = ["NVDA", "TSLA", "TSM", "ANET", "AMD", "INTC", "AVGO", "MSFT", "ORCL", "MU", "GOOG", "GOOGL",
           "CRCL", "LITE", "COHR", "QQQ", "SPY", "SMH", "SOXX", "IBB", "AAPL", "AMZN", "META", "NFLX"]

# คอลัมน์: (header, key, width, number_format, dropdown_inline|None)
COLS = {
    "客戶": [
        ("姓名 *", "name", 18, None, None),
        ("美元額度", "usd_amount", 15, NUM, None),
        ("幣別", "currency", 10, None, CURRENCIES),
        ("統一帳號", "unified_account", 16, None, None),
        ("備註", "notes", 34, None, None),
    ],
    "商品": [
        ("商品代號 *", "product_code", 16, None, None),
        ("類別", "category", 12, None, CATEGORIES),
        ("標的1", "underlying_1", 11, None, TICKERS),
        ("期初價1", "initial_price_1", 11, NUM2, None),
        ("標的2", "underlying_2", 11, None, TICKERS),
        ("期初價2", "initial_price_2", 11, NUM2, None),
        ("標的3", "underlying_3", 11, None, TICKERS),
        ("期初價3", "initial_price_3", 11, NUM2, None),
        ("KO障壁(%)", "ko_barrier", 11, PCT, None),
        ("KI障壁(%)", "ki_barrier", 11, PCT, None),
        ("履約價(%)", "strike_pct", 11, PCT, None),
        ("票息(%年化)", "coupon_pct", 12, PCT, None),
        ("配息頻率", "coupon_freq", 12, None, FREQ),
        ("成交日", "trade_date", 14, DATE, None),
        ("比價日", "observation_date", 14, DATE, None),
        ("到期日", "exit_date", 14, DATE, None),
        ("狀態", "status", 11, None, STATUS),
    ],
    "投資": [
        ("客戶姓名 *", "customer", 18, None, "ref:InvestorList"),
        ("商品代號 *", "product_code", 16, None, "ref:ProductList"),
        ("投資金額 *", "amount_usd", 15, NUM, None),
        ("幣別", "currency", 10, None, CURRENCIES),
    ],
}
SUBTITLE = {
    "客戶": "投資人名單 · 必填：姓名",
    "商品": "結構型商品 · 必填：商品代號 · % 欄填數字 (KO=100…)",
    "投資": "持倉 · 客戶姓名/商品代號 可從下拉選 (對應其他分頁)",
}


def _thin():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)


def _build_sheet(ws, title, rows):
    border = _thin()
    ws.sheet_view.showGridLines = False
    cols = COLS[title]
    ncol = len(cols)
    last = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = f"  Justinvestment · {title}"
    ws["A1"].fill = PatternFill("solid", fgColor=GREEN)
    ws["A1"].font = Font(name=FONT, bold=True, color=WHITE, size=16)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = "  " + SUBTITLE[title]
    ws["A2"].fill = PatternFill("solid", fgColor=TINT)
    ws["A2"].font = Font(name=FONT, color=GREEN_DK, size=10.5)
    ws.row_dimensions[2].height = 22

    HR = 3
    for i, (h, _k, w, _nf, _dd) in enumerate(cols, 1):
        c = ws.cell(HR, i, h)
        c.fill = PatternFill("solid", fgColor=GREEN_DK)
        c.font = Font(name=FONT, bold=True, color=WHITE, size=11.5)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HR].height = 30
    ws.freeze_panes = f"A{HR+1}"

    data = rows or []
    total = len(data) + BLANK_ROWS
    for ridx in range(total):
        r = HR + 1 + ridx
        rec = data[ridx] if ridx < len(data) else {}
        fill = PatternFill("solid", fgColor=WHITE if r % 2 == 0 else ZEBRA)
        for i, (_h, k, _w, nf, _dd) in enumerate(cols, 1):
            c = ws.cell(r, i, rec.get(k) if rec else None)
            c.fill = fill; c.border = border
            c.font = Font(name=FONT, size=11.5, color=INK)
            c.alignment = Alignment(vertical="center")
            if nf:
                c.number_format = nf
        ws.row_dimensions[r].height = 22

    # validations
    end = HR + total
    for i, (h, _k, _w, nf, dd) in enumerate(cols, 1):
        col = get_column_letter(i)
        rng = f"{col}{HR+1}:{col}{end}"
        if isinstance(dd, list):
            dv = DataValidation(type="list", formula1='"' + ",".join(dd) + '"', allow_blank=True)
            dv.showErrorMessage = (dd not in (TICKERS,))  # ticker = loose (allow other)
            ws.add_data_validation(dv); dv.add(rng)
        elif isinstance(dd, str) and dd.startswith("ref:"):
            dv = DataValidation(type="list", formula1=dd[4:], allow_blank=True)
            dv.showInputMessage = True; dv.prompt = "可從下拉選擇"
            ws.add_data_validation(dv); dv.add(rng)
        elif nf == DATE:
            dv = DataValidation(type="date", operator="between", formula1="2000-01-01", formula2="2100-12-31", allow_blank=True)
            dv.showInputMessage = True; dv.prompt = "日期 YYYY-MM-DD"
            ws.add_data_validation(dv); dv.add(rng)
    return end


def build_workbook(data: dict = None, with_sample: bool = False) -> Workbook:
    from openpyxl.workbook.defined_name import DefinedName
    data = data or {}
    cust = list(data.get("customers") or [])
    prod = list(data.get("products") or [])
    inv = list(data.get("investments") or [])
    if with_sample and not (cust or prod or inv):
        cust = [{"name": "蔣太太", "usd_amount": 370000, "currency": "USD", "notes": "← 範例，請刪除"}]
        prod = [{"product_code": "EQDS0702653", "category": "SN", "underlying_1": "TSLA", "initial_price_1": 250.5,
                 "underlying_2": "TSM", "initial_price_2": 180.2, "ko_barrier": 100, "ki_barrier": 60,
                 "strike_pct": 100, "coupon_pct": 8, "coupon_freq": "月配",
                 "trade_date": "2026-05-11", "observation_date": "2026-06-18", "status": "進行中"}]
        inv = [{"customer": "蔣太太", "product_code": "EQDS0702653", "amount_usd": 370000, "currency": "USD"}]

    wb = Workbook(); wb.remove(wb.active)

    # 說明
    ws0 = wb.create_sheet("說明"); ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 4; ws0.column_dimensions["B"].width = 96
    ws0.merge_cells("A1:B2")
    ws0["A1"] = "  Justinvestment · 資料表"
    ws0["A1"].fill = PatternFill("solid", fgColor=GREEN); ws0["A1"].font = Font(name=FONT, bold=True, color=WHITE, size=18)
    ws0["A1"].alignment = Alignment(vertical="center"); ws0.row_dimensions[1].height = 30
    notes = [
        "", "在「客戶 / 商品 / 投資」分頁輸入或續填資料，版面已排好。",
        "已有的資料會顯示在上方，往下接著填新的即可，存檔後可再匯入系統。",
        "日期 YYYY-MM-DD · 百分比填數字 (KO=100、票息=8) · 配息頻率/狀態可下拉選。",
        "「投資」分頁的客戶姓名/商品代號可從下拉選 (對應其他分頁)。",
    ]
    for i, txt in enumerate(notes, 4):
        cc = ws0.cell(i, 2, txt); cc.font = Font(name=FONT, size=11, color=INK); cc.alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[i].height = 22 if txt else 8

    ws1 = wb.create_sheet("客戶"); end1 = _build_sheet(ws1, "客戶", cust)
    ws2 = wb.create_sheet("商品"); _build_sheet(ws2, "商品", prod)
    ws3 = wb.create_sheet("投資"); end3 = _build_sheet(ws3, "投資", inv)

    # defined names for 投資 reference dropdowns (cover all rows)
    dnI = DefinedName(name="InvestorList", attr_text=f"'客戶'!$A$4:$A${end1}")
    dnP = DefinedName(name="ProductList", attr_text=f"'商品'!$A$4:$A${end1}")
    try:
        wb.defined_names["InvestorList"] = dnI
        wb.defined_names["ProductList"] = dnP
    except TypeError:
        wb.defined_names.add(dnI); wb.defined_names.add(dnP)
    _ = end3
    return wb


# ── 統一證券 庫存查詢 ฟอร์แมต (แถวละ 1 持倉, 32 คอลัมน์) — export ตรง Template.xlsx, round-trip import ได้ ──
UNIDOS_HEADERS = [
    "PSC商品代碼", "幣別", "名目本金(計價幣)", "客戶名稱", "交易日", "最後保證配息日", "發行日",
    "期末訂價日", "商品到期日", "產品別", "連結標的1", "連結標的2", "連結標的3", "連結標的4", "連結標的5",
    "天期 (月)", "利率", "配息頻率", "配息區間/執行價", "提前出場型式", "提前出場價", "下限型式", "界限價",
    "交割日", "保證配息月數", "成交上手", "價格", "期初價格1", "期初價格2", "期初價格3", "期初價格4", "期初價格5",
]


def build_unidos_workbook(rows: list) -> Workbook:
    """rows = list[dict] keyed by UNIDOS_HEADERS → sheet 庫存查詢 (統一證券 ฟอร์แมต)。
    มีกรอบเส้น + แถบสลับสี + freeze หัวตาราง เพื่อให้อ่านง่าย。"""
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("庫存查詢")
    ws.sheet_view.showGridLines = False
    border = _thin()
    ncol = len(UNIDOS_HEADERS)

    ws.append(UNIDOS_HEADERS)
    head_fill = PatternFill("solid", fgColor=GREEN_DK)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = max(11, len(str(UNIDOS_HEADERS[c - 1])) + 2)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for ridx, r in enumerate(rows):
        excel_row = ridx + 2  # +1 header, +1 1-based
        ws.append([r.get(h) for h in UNIDOS_HEADERS])
        fill = PatternFill("solid", fgColor=WHITE if excel_row % 2 == 0 else ZEBRA)
        for c in range(1, ncol + 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name=FONT, size=10.5, color=INK)
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[excel_row].height = 20
    return wb
