"""
Excel Parser - ใช้งานร่วมกันระหว่าง Web Upload และ Script
รองรับทุกเดือน, หลายไฟล์, หลาย sheet
"""
import re
import openpyxl
from datetime import datetime
from typing import Union, Optional
from io import BytesIO


def detect_month_label(sheet_name: str) -> str:
    """ตรวจจับชื่อเดือนจากชื่อ sheet"""
    fullwidth_digits = {'１':'1','２':'2','３':'3','４':'4','５':'5',
                        '６':'6','７':'7','８':'8','９':'9','０':'0'}
    normalized = ''.join(fullwidth_digits.get(c, c) for c in sheet_name)
    match = re.search(r'(\d+)月', normalized)
    if match:
        return f"{match.group(1)}月"
    chinese_months = {
        '一':'1','二':'2','三':'3','四':'4','五':'5',
        '六':'6','七':'7','八':'8','九':'9','十':'10',
        '十一':'11','十二':'12'
    }
    for cn, num in chinese_months.items():
        if cn + '月' in sheet_name:
            return f"{num}月"
    return sheet_name.strip()


def is_customer_sheet(sheet_name: str) -> bool:
    """ตรวจว่าเป็น sheet ลูกค้าไหม"""
    keywords = ['開戶', '客戶', '戶名', 'customer', 'CLIENT']
    return any(k.lower() in sheet_name.lower() for k in keywords)


def is_sn_sheet(sheet_name: str) -> bool:
    """ตรวจว่าเป็น sheet SN ไหม"""
    name = sheet_name.upper()
    # Full-width ＳＮ or regular SN
    return ('ＳＮ' in sheet_name or 'SN' in name) and '月' in sheet_name


def normalize_pct(value) -> Optional[float]:
    """ทำให้เป็น decimal (0.80 = 80%)"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        if v == 0:
            return None
        if v > 5:       # 74.87 → 0.7487
            return round(v / 100, 6)
        return round(v, 6)
    return None


def normalize_bool(value) -> bool:
    if value is None:
        return False
    return str(value).strip().upper() in ['V', 'Ｖ', 'TRUE', '1', 'Y']


def clean_code(code) -> str:
    """ทำความสะอาดรหัสสินค้า"""
    if not code:
        return ""
    code = str(code).strip()
    fullwidth = {'Ａ':'A','Ｂ':'B','Ｃ':'C','Ｄ':'D','Ｅ':'E','Ｆ':'F',
                 'Ｇ':'G','Ｈ':'H','Ｉ':'I','Ｊ':'J','Ｋ':'K','Ｌ':'L',
                 'Ｍ':'M','Ｎ':'N','Ｏ':'O','Ｐ':'P','Ｑ':'Q','Ｒ':'R',
                 'Ｓ':'S','Ｔ':'T','Ｕ':'U','Ｖ':'V','Ｗ':'W','Ｘ':'X',
                 'Ｙ':'Y','Ｚ':'Z','１':'1','２':'2','３':'3','４':'4',
                 '５':'5','６':'6','７':'7','８':'8','９':'9','０':'0'}
    return ''.join(fullwidth.get(c, c) for c in code)


# ============================================================
# Parse 客戶 Sheet
# ============================================================

def parse_customer_sheet(ws) -> list:
    """Parse ลูกค้าจาก sheet → return list of dict"""
    customers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if not name or name in ['戶名', '姓名']:
            continue

        customers.append({
            "name": name,
            "unified_account": normalize_bool(row[1] if len(row) > 1 else None),
            "pi_signed":       normalize_bool(row[2] if len(row) > 2 else None),
            "ordered":         normalize_bool(row[3] if len(row) > 3 else None),
            "usd_amount":      float(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else None,
            "ctbc_position":   float(row[5]) if len(row) > 5 and isinstance(row[5], (int, float)) else None,
            "fund_amount":     float(row[6]) if len(row) > 6 and isinstance(row[6], (int, float)) else None,
        })
    return customers


# ============================================================
# Parse SN Sheet
# ============================================================

def parse_sn_sheet(ws, month_label: str) -> list:
    """
    Parse SN sheet → return list of SN dicts (each with 'investments' list)

    รูปแบบ Excel:
    - แถว SN: วันที่ในคอลัมน์ A, รหัสในคอลัมน์ B
    - แถวราคาเริ่มต้น: 期初價格 ในคอลัมน์ B (หรืออยู่ในแถวลูกค้า)
    - แถวลูกค้า: ชื่อในคอลัมน์ A, จำนวนเงินในคอลัมน์ B
    """
    rows = list(ws.iter_rows(values_only=True))
    sn_list = []
    current_sn = None

    i = 0
    while i < len(rows):
        row = rows[i]

        # --- ตรวจ SN header row (คอลัมน์ A เป็น datetime, คอลัมน์ B เป็นรหัส) ---
        col_a = row[0]
        col_b = row[1] if len(row) > 1 else None

        is_sn_row = (
            isinstance(col_a, datetime) and
            col_b and
            isinstance(col_b, str) and
            col_b.strip() and
            col_b.strip() != '期初價格'
        )

        if is_sn_row:
            if current_sn:
                sn_list.append(current_sn)

            raw_code = str(col_b).strip()
            product_code = clean_code(raw_code)

            # Underlyings (อาจอยู่ในคอลัมน์ C-G)
            underlyings = []
            for j in range(2, 7):
                val = row[j] if len(row) > j else None
                if val and isinstance(val, str) and val.strip():
                    underlyings.append(val.strip().upper())
                else:
                    underlyings.append(None)

            current_sn = {
                "product_code": product_code,
                "trade_date": col_a.strftime('%Y-%m-%d'),
                "underlying_1": underlyings[0],
                "underlying_2": underlyings[1],
                "underlying_3": underlyings[2],
                "underlying_4": underlyings[3],
                "underlying_5": underlyings[4],
                "initial_price_1": None,
                "initial_price_2": None,
                "initial_price_3": None,
                "initial_price_4": None,
                "initial_price_5": None,
                "strike_pct":       normalize_pct(row[7] if len(row) > 7 else None),
                "coupon_pct":       normalize_pct(row[8] if len(row) > 8 else None),
                "observation_date": row[9].strftime('%Y-%m-%d') if len(row) > 9 and isinstance(row[9], datetime) else (str(row[9])[:10] if len(row) > 9 and row[9] else None),
                "ko_barrier":       normalize_pct(row[10] if len(row) > 10 else None),
                "ki_barrier":       normalize_pct(row[11] if len(row) > 11 else None),
                "exit_date":        row[12].strftime('%Y-%m-%d') if len(row) > 12 and isinstance(row[12], datetime) else None,
                "temp_settlement":  float(row[13]) if len(row) > 13 and isinstance(row[13], (int, float)) else None,
                "chu":              str(row[14]).strip() if len(row) > 14 and row[14] else None,
                "total_order_amount": float(row[15]) if len(row) > 15 and isinstance(row[15], (int, float)) else None,
                "month_label":      month_label,
                "status":           "active",
                "investments":      [],
            }
            i += 1
            continue

        # --- แถวราคาเริ่มต้น: มีป้ายกำกับ "期初價格" หรือ col_a/col_b ว่าง + col_c เป็นตัวเลข ---
        is_labeled   = col_b and '期初' in str(col_b)
        is_unlabeled = (col_a is None and col_b is None and
                        len(row) > 2 and isinstance(row[2], (int, float)) and float(row[2]) > 0)
        if current_sn and (is_labeled or is_unlabeled):
            for j in range(5):
                val = row[j + 2] if len(row) > j + 2 else None
                current_sn[f"initial_price_{j+1}"] = float(val) if isinstance(val, (int, float)) else None
            i += 1
            continue

        # --- แถวลูกค้า (คอลัมน์ A ชื่อ, คอลัมน์ B จำนวนเงิน) ---
        if (current_sn and
                col_a and isinstance(col_a, str) and col_a.strip() and
                isinstance(col_b, (int, float))):

            customer_name = col_a.strip()
            amount = float(col_b)

            # บางครั้งราคาเริ่มต้นอยู่ในแถวลูกค้า (ไม่มีแถว 期初價格 แยก)
            if current_sn["initial_price_1"] is None:
                for j in range(5):
                    val = row[j + 2] if len(row) > j + 2 else None
                    if isinstance(val, (int, float)) and val > 0:
                        current_sn[f"initial_price_{j+1}"] = float(val)

            current_sn["investments"].append({
                "customer_name": customer_name,
                "amount_usd": amount,
            })
            i += 1
            continue

        i += 1

    if current_sn:
        sn_list.append(current_sn)

    return sn_list


# ============================================================
# Main: Parse ไฟล์ Excel ทั้งหมด
# ============================================================

def parse_excel_file(file_source: Union[str, BytesIO]) -> dict:
    """
    Parse ไฟล์ Excel → return dict พร้อม import

    Returns:
        {
            "customers": [...],
            "sn_by_month": {
                "5月": [...],
                "6月": [...],
            },
            "sheets_found": {
                "customer_sheets": [...],
                "sn_sheets": [...],
                "other_sheets": [...],
            }
        }
    """
    if isinstance(file_source, str):
        wb = openpyxl.load_workbook(file_source)
    else:
        wb = openpyxl.load_workbook(file_source)

    result = {
        "customers": [],
        "sn_by_month": {},
        "sheets_found": {"customer_sheets": [], "sn_sheets": [], "other_sheets": []}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if is_customer_sheet(sheet_name):
            result["sheets_found"]["customer_sheets"].append(sheet_name)
            customers = parse_customer_sheet(ws)
            result["customers"].extend(customers)

        elif is_sn_sheet(sheet_name):
            month = detect_month_label(sheet_name)
            result["sheets_found"]["sn_sheets"].append(sheet_name)
            sns = parse_sn_sheet(ws, month)
            if month not in result["sn_by_month"]:
                result["sn_by_month"][month] = []
            result["sn_by_month"][month].extend(sns)

        else:
            result["sheets_found"]["other_sheets"].append(sheet_name)

    return result


def get_summary(parsed: dict) -> dict:
    """สรุปข้อมูลที่ parse ได้"""
    total_sns = sum(len(v) for v in parsed["sn_by_month"].values())
    total_investments = sum(
        len(sn["investments"])
        for month_sns in parsed["sn_by_month"].values()
        for sn in month_sns
    )
    return {
        "customers": len(parsed["customers"]),
        "months": list(parsed["sn_by_month"].keys()),
        "total_sns": total_sns,
        "total_investments": total_investments,
    }
