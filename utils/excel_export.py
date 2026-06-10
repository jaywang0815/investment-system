"""
Excel 匯出 & Google Sheets 同步模組
"""
import io
from datetime import date
from typing import Optional
import pandas as pd


# ============================================================
# Shared B-style Excel builder (used by both LINE bot and web)
# ============================================================

def build_excel_bytes(customers: list = None, sns_all: list = None, sn_inv_map: dict = None,
                      source_bytes: bytes = None) -> bytes:
    """
    Style the source Excel (assets/source_data.xlsx) and return bytes.
    source_bytes: optional raw bytes of the source file (e.g. uploaded by user).
    customers / sns_all / sn_inv_map: unused — kept for backwards-compat signature.

    Layout rules applied to the source:
    - Title row 1: remove any existing title, add navy bar "客戶開戶明細" / "ＳＮ{月} 商品明細"
    - Date row 2: navy subtitle with today's date
    - Header row 3 (source row 1): navy2 bg, white bold text
    - Data rows 4+ (source rows 2+): alternating GRAY1/WHITE bg
      - Customer names (col A): dark near-black, bold, size 11
      - 代號 (SN sheet col B) + 標的1-5 (cols C-G) on SN product rows: amber bold
      - V marks: green
    - Date columns (日期, 比價): YYYY-MM-DD format, col width ≥ 14
    """
    import os
    from copy import copy as _copy
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    from utils import branding as B
    SRC_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets", "source_data.xlsx")
    NAVY = B.C_PRIMARY; NAVY2 = B.C_HEADER; WHITE = "FFFFFF"
    GREEN = "1B9E5A"; GRAY1 = B.C_ZEBRA; DARK = "1A1A1A"; AMBER = "B45309"
    DATEBG = "7E2A1F"   # 日期副標底 (深紅)
    today_str = date.today().strftime("%Y年%m月%d日") + "　　" + B.SIGNATURE

    def bdr(c="CBD5E1"):
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    def apply(cell, bold=False, align="center", bg=WHITE, color=DARK, sz=10, fmt=None):
        cell.font      = Font(name="微軟正黑體", bold=bold, color=color, size=sz)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = bdr()
        if fmt:
            cell.number_format = fmt

    # ── Load source file (data untouched) ────────────────────────────
    if source_bytes:
        src_wb = load_workbook(io.BytesIO(source_bytes))
    else:
        src_wb = load_workbook(SRC_PATH)

    out_wb = __import__("openpyxl").Workbook()
    first  = True

    SHEET_TITLES = {
        "開戶明細": "客戶開戶明細",
        "ＳＮ5月":  "ＳＮ5月 商品明細",
    }

    # SN sheet: which col indices are 代號(2) + 標的1-5(3-7)
    SN_AMBER = {2, 3, 4, 5, 6, 7}
    # SN sheet: date columns → YYYY-MM-DD
    SN_DATE  = {1, 10, 13}
    # SN sheet col widths
    SN_W = {"A":14,"B":18,"C":12,"D":12,"E":12,"F":12,"G":12,
            "H":10,"I":10,"J":14,"K":10,"L":12,"M":12,"N":12,"O":10,"P":14}

    for src_name in src_wb.sheetnames:
        src_ws   = src_wb[src_name]
        is_sn    = (src_name != "開戶明細")
        out_name = ("客戶總覽" if src_name == "開戶明細"
                    else "SN明細_" + src_name.replace("ＳＮ", "").replace("SN", "").strip())
        title    = SHEET_TITLES.get(src_name, src_name)

        ws = out_wb.active if first else out_wb.create_sheet()
        first = False
        ws.title = out_name
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

        max_r = src_ws.max_row
        max_c = src_ws.max_column

        # ── Row 1: title bar ─────────────────────────────────────────
        ws.row_dimensions[1].height = 40
        ws.merge_cells(f"A1:{get_column_letter(max_c)}1")
        t = ws.cell(1, 1, value=title)
        t.font      = Font(name="微軟正黑體", bold=True, color=WHITE, size=14)
        t.fill      = PatternFill("solid", fgColor=NAVY)
        t.alignment = Alignment(horizontal="left", vertical="center", indent=2)

        # logo (top-right corner, floats over title bar)
        try:
            if B.has_logo():
                from openpyxl.drawing.image import Image as XLImage
                logo = XLImage(B.LOGO_PATH)
                logo.width = 46; logo.height = 46
                ws.add_image(logo, f"{get_column_letter(max_c)}1")
        except Exception:
            pass

        # ── Row 2: date subtitle + 報告人署名 ─────────────────────────
        ws.row_dimensions[2].height = 20
        ws.merge_cells(f"A2:{get_column_letter(max_c)}2")
        s = ws.cell(2, 1, value=today_str)
        s.font      = Font(name="微軟正黑體", color="F4D9D3", size=9)
        s.fill      = PatternFill("solid", fgColor=DATEBG)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # ── Row 3: source row 1 = column headers ─────────────────────
        ws.row_dimensions[3].height = 28
        for c in range(1, max_c + 1):
            val  = src_ws.cell(1, c).value
            cell = ws.cell(3, c, value=val)
            cell.font      = Font(name="微軟正黑體", bold=True, color=WHITE, size=11)
            cell.fill      = PatternFill("solid", fgColor=NAVY2)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr("334155")

        # ── Rows 4+: source rows 2+ = data (values unchanged) ────────
        # Detect SN product rows: col A = datetime
        from datetime import datetime as _dt
        for sr in range(2, max_r + 1):
            out_r = sr + 2
            ws.row_dimensions[out_r].height = 20
            bg = GRAY1 if out_r % 2 == 0 else WHITE

            col_a = src_ws.cell(sr, 1).value
            col_b = src_ws.cell(sr, 2).value
            is_sn_prod_row = is_sn and isinstance(col_a, _dt)

            for c in range(1, max_c + 1):
                src_cell = src_ws.cell(sr, c)
                val      = src_cell.value
                cell     = ws.cell(out_r, c, value=val)
                cell.number_format = src_cell.number_format

                if is_sn_prod_row and c in SN_AMBER and isinstance(val, str) and val.strip():
                    # amber bold: 代號 and 標的 on SN product rows
                    apply(cell, bold=True, align="left" if c == 2 else "center",
                          bg=bg, color=AMBER, sz=11)

                elif c == 1 and isinstance(val, str) and val.strip():
                    # customer name (col A, string) → dark bold
                    apply(cell, bold=True, align="left", bg=bg, color=DARK, sz=11)

                else:
                    is_v = isinstance(val, str) and val.strip() in ("V", "Ｖ", "v", "X", "Ｘ")
                    apply(cell, bold=is_v,
                          align="left" if c == 1 else "center",
                          bg=bg, color=GREEN if (is_v and val.strip() not in ("X","Ｘ")) else DARK)

                # date format for date-type cells (no time)
                if is_sn and c in SN_DATE and isinstance(val, _dt):
                    cell.number_format = "YYYY-MM-DD"
                # numeric format for amounts
                if is_sn and c == 16 and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

        # column widths
        if is_sn:
            for col_letter, w in SN_W.items():
                ws.column_dimensions[col_letter].width = w
        else:
            cust_widths = {"A":20,"B":12,"C":12,"D":12,"E":16,"F":16,"G":16,"H":16}
            for col_letter, w in cust_widths.items():
                ws.column_dimensions[col_letter].width = w
            for ci in range(9, max_c + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 14

        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue()

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False


# ============================================================
# Excel 匯出
# ============================================================

def export_to_excel(customers_df: pd.DataFrame, sns_df: pd.DataFrame,
                    investments_df: pd.DataFrame) -> bytes:
    """匯出所有資料為 Excel 檔案"""
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # 客戶資料
        customers_out = customers_df.copy()
        for col in ["unified_account", "pi_signed", "ordered"]:
            if col in customers_out.columns:
                customers_out[col] = customers_out[col].map({True: "Ｖ", False: ""})
        col_rename = {
            "name": "戶名", "unified_account": "統一開戶",
            "pi_signed": "PI見簽", "ordered": "已下單",
            "usd_amount": "USD金額", "ctbc_position": "中信部位",
            "fund_amount": "FUND", "notes": "備註"
        }
        customers_out = customers_out.rename(columns=col_rename)
        display_cols = [c for c in col_rename.values() if c in customers_out.columns]
        customers_out[display_cols].to_excel(writer, sheet_name="客戶資料", index=False)

        # SN 商品資料
        sns_out = sns_df.copy()
        sn_col_rename = {
            "product_code": "商品代號", "trade_date": "交易日期",
            "underlying_1": "標的1", "underlying_2": "標的2",
            "underlying_3": "標的3", "underlying_4": "標的4",
            "underlying_5": "標的5", "initial_price_1": "期初價1",
            "initial_price_2": "期初價2", "initial_price_3": "期初價3",
            "initial_price_4": "期初價4", "initial_price_5": "期初價5",
            "strike_pct": "執行價%", "coupon_pct": "配息%",
            "observation_date": "比價日", "ko_barrier": "KO水位",
            "ki_barrier": "KI水位", "status": "狀態",
            "total_order_amount": "總下單金額", "month_label": "月份"
        }
        sns_out = sns_out.rename(columns=sn_col_rename)
        for col in ["執行價%", "配息%", "KO水位", "KI水位"]:
            if col in sns_out.columns:
                sns_out[col] = sns_out[col].apply(
                    lambda x: f"{x*100:.2f}%" if pd.notna(x) else ""
                )
        status_map = {
            "active": "有效", "ko_triggered": "KO觸發",
            "ki_triggered": "KI觸發", "expired": "已到期", "matured": "已結算"
        }
        if "狀態" in sns_out.columns:
            sns_out["狀態"] = sns_out["狀態"].map(status_map).fillna(sns_out["狀態"])
        display_sn_cols = [c for c in sn_col_rename.values() if c in sns_out.columns]
        sns_out[display_sn_cols].to_excel(writer, sheet_name="SN商品", index=False)

        # 投資記錄
        if not investments_df.empty:
            inv_out = investments_df.copy()
            inv_col_rename = {
                "customer_name": "客戶姓名", "product_code": "商品代號",
                "amount_usd": "投資金額(USD)", "underlying_1": "標的1",
                "underlying_2": "標的2", "underlying_3": "標的3",
                "observation_date": "比價日", "status": "狀態"
            }
            inv_out = inv_out.rename(columns=inv_col_rename)
            display_inv_cols = [c for c in inv_col_rename.values() if c in inv_out.columns]
            inv_out[display_inv_cols].to_excel(writer, sheet_name="投資記錄", index=False)

        # 格式化
        wb = writer.book
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    return buffer.getvalue()


# ============================================================
# Google Sheets 同步
# ============================================================

def sync_to_google_sheets(customers_df: pd.DataFrame, sns_df: pd.DataFrame,
                           sheet_id: Optional[str] = None) -> bool:
    """同步資料到 Google Sheets"""
    import streamlit as st
    if not GSPREAD_AVAILABLE:
        st.error("gspread 未安裝，請執行 pip install gspread")
        return False

    try:
        sheet_id = sheet_id or st.secrets.get("GOOGLE_SHEET_ID")
        if not sheet_id:
            st.error("請在 secrets.toml 設定 GOOGLE_SHEET_ID")
            return False

        # 使用 service account
        creds_info = st.secrets.get("GOOGLE_SERVICE_ACCOUNT")
        if not creds_info:
            st.error("請在 secrets.toml 設定 GOOGLE_SERVICE_ACCOUNT")
            return False

        creds = Credentials.from_service_account_info(
            dict(creds_info),
            scopes=["https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"]
        )
        gc = gspread.authorize(creds)
        spreadsheet = gc.open_by_key(sheet_id)

        # 同步客戶資料
        try:
            ws = spreadsheet.worksheet("客戶資料")
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet("客戶資料", rows=200, cols=20)

        ws.clear()
        customers_out = customers_df[["name", "usd_amount", "ctbc_position",
                                       "unified_account", "pi_signed", "ordered"]].copy()
        customers_out.columns = ["戶名", "USD金額", "中信部位", "統一開戶", "PI見簽", "已下單"]
        ws.update([customers_out.columns.tolist()] + customers_out.fillna("").values.tolist())

        # 同步 SN 商品
        try:
            ws2 = spreadsheet.worksheet("SN商品")
        except gspread.WorksheetNotFound:
            ws2 = spreadsheet.add_worksheet("SN商品", rows=200, cols=20)

        ws2.clear()
        sns_out = sns_df[["product_code", "trade_date", "underlying_1", "underlying_2",
                            "underlying_3", "strike_pct", "coupon_pct",
                            "observation_date", "ko_barrier", "ki_barrier", "status"]].copy()
        sns_out.columns = ["代號", "日期", "標的1", "標的2", "標的3",
                            "執行價%", "配息%", "比價日", "KO水位", "KI水位", "狀態"]
        ws2.update([sns_out.columns.tolist()] + sns_out.fillna("").astype(str).values.tolist())

        return True

    except Exception as e:
        st.error(f"Google Sheets 同步失敗: {e}")
        return False
