"""資料匯入 — 讀標準範本 (依「標題」非位置)，預覽後確認，寫入登入者自己的 tenant。
無 AI、不外傳資料；% 欄 ÷100；日期 ISO；配息頻率/狀態 中→英 對應。"""
import io
import math
import re as _re
import unicodedata
from datetime import datetime, date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from ..deps import repo
from ..db import Repo

router = APIRouter(prefix="/api/import", tags=["import"])


def _safe_bulk(r: Repo, table: str, rows: list):
    """bulk insert + เติม tenant_id + ตัดคอลัมน์ที่ DB ยังไม่มีแล้วลองใหม่ (กัน migration ไม่ครบ)。คืนแถวที่สร้าง。"""
    if not rows:
        return []
    payload = [{**row, "tenant_id": r.tenant_id} for row in rows]
    for _ in range(8):
        try:
            return r.sb.table(table).insert(payload).execute().data or []
        except Exception as e:
            msg = getattr(e, "message", None) or str(e)
            mm = (_re.search(r"column \S*?\.?(\w+) does not exist", msg)
                  or _re.search(r"Could not find the '(\w+)' column", msg))
            if not mm:
                raise
            bad, removed = mm.group(1), False
            for row in payload:
                if bad in row:
                    row.pop(bad, None); removed = True
            if not removed:
                raise
    return r.sb.table(table).insert(payload).execute().data or []


def _norm(s) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return s.replace("*", "").replace(" ", "").replace("\n", "").strip().lower()


# field -> header aliases (normalized compare)
CUST = {
    "name": ["姓名", "戶名", "客戶姓名", "name", "客戶"],
    "usd_amount": ["美元額度", "usd", "美金", "美元"],
    "unified_account": ["統一開戶", "統一帳號"],   # BOOLEAN (Ｖ → true)
    "pi_signed": ["pi見簽", "見簽", "pi"],          # BOOLEAN
    "ordered": ["已下單"],                          # BOOLEAN
    "ctbc_position": ["中信部位", "中信"],
    "fund_amount": ["fund", "基金部位"],
    "notes": ["備註", "備注", "notes", "note"],
}
PROD = {
    "product_code": ["商品代號", "代號", "code", "product"],
    "category": ["類別", "category", "類型"],
    "underlying_1": ["標的1", "標的一", "underlying1", "標的"],
    "initial_price_1": ["期初價1", "期初價格1", "初始價1", "initial1"],
    "underlying_2": ["標的2", "標的二", "underlying2"],
    "initial_price_2": ["期初價2", "期初價格2", "initial2"],
    "underlying_3": ["標的3", "標的三", "underlying3"],
    "initial_price_3": ["期初價3", "期初價格3", "initial3"],
    "ko_barrier": ["ko障壁", "ko水位", "ko提前", "ko"],
    "ki_barrier": ["ki障壁", "ki水位", "ki下限", "ki"],
    "strike_pct": ["履約價", "執行價", "strike", "履約"],
    "coupon_pct": ["票息", "配息", "coupon"],
    "coupon_freq": ["配息頻率", "頻率", "freq"],
    "trade_date": ["成交日", "交易日", "trade"],
    "observation_date": ["比價日", "比價", "觀察日", "obs"],
    "exit_date": ["到期日", "出場日", "到期", "exit", "maturity"],
    "status": ["狀態", "status"],
}
INV = {
    "customer": ["客戶姓名", "姓名", "戶名", "客戶", "customer"],
    "product_code": ["商品代號", "代號", "code"],
    "amount_usd": ["投資金額", "金額", "amount", "下單金額"],
    "currency": ["幣別", "currency"],
}
FREQ_MAP = {"月配": "monthly", "季配": "quarterly", "半年配": "semiannual", "年配": "annual", "到期一次": "maturity"}
STATUS_MAP = {"進行中": "active", "已出場": "exited", "暫停": "inactive"}
SHEET_HINT = {"客戶": "cust", "投資人": "cust", "customer": "cust", "開戶": "cust",
              "商品": "prod", "product": "prod", "sn": "prod",
              "投資": "inv", "持倉": "inv", "holding": "inv"}

# 「區塊式」SN 表頭 (DOUU 現用格式)：每檔商品一列，下方接「期初價格」列與投資人列
SN_BLOCK_HDR = {
    "date": ["日期", "成交日", "date"],
    "code": ["代號", "商品代號", "code"],
    "tenor": ["期間", "年期", "期限", "tenor"],
    "u1": ["標的1", "標的一"],
    "u2": ["標的2", "標的二"],
    "u3": ["標的3", "標的三"],
    "u4": ["標的4", "標的四"],
    "u5": ["標的5", "標的五"],
    "strike": ["執行價", "履約價", "strike"],
    "coupon": ["配息", "票息", "coupon"],
    "obs": ["比價", "比價日", "觀察日", "obs"],
    "ko": ["ko提前", "ko障壁", "ko水位", "ko"],
    "ki": ["ki下限", "ki障壁", "ki水位", "ki"],
    "exit": ["出場", "到期日", "exit", "maturity"],
}

# 「券商庫存匯出」格式 (扁平，每列一筆持倉=商品+客戶+金額)；標的含交易所後綴 "TSLA UW"
# 配息起算(比價日) = 最後保證配息日 (≈ปลายเดือน, advisor ยืนยัน)；配息頻率=月配
BROKER_HDR = {
    "code": ["psc商品代碼", "商品代碼"],
    "currency": ["幣別"],
    "amount": ["名目本金"],
    "customer": ["客戶名稱"],
    "trade_date": ["交易日"],
    "obs_date": ["最後保證配息日"],   # = วันจ่ายดอกงวดแรก/保證 (≈ปลายเดือน) → ดอกเริ่มนับจากนี่ (advisor ยืนยัน "6月底")
    "exit_date": ["商品到期日"],
    "ptype": ["產品別"],
    "u1": ["連結標的1"], "u2": ["連結標的2"], "u3": ["連結標的3"], "u4": ["連結標的4"], "u5": ["連結標的5"],
    "coupon": ["利率"],
    "strike": ["執行價"],
    "ko": ["提前出場價"],
    "ki": ["界限價"],
    "ip1": ["期初價格1"], "ip2": ["期初價格2"], "ip3": ["期初價格3"], "ip4": ["期初價格4"], "ip5": ["期初價格5"],
}


def _year_of(v):
    """ดึงปีจากค่าวันที่ (datetime/date/ROC-int) เพื่อใช้เป็นปีให้กับ "M月D日" ที่ไม่มีปี。"""
    if isinstance(v, (datetime, date)):
        return v.year
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        iv = int(v)
        if 1000000 <= iv <= 9991231:
            y = iv // 10000
            return y + 1911 if y < 1911 else y
    return None


def _to_date(v, year_hint=None):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    # 民國年整數 (ROC) 例如 1150608 → 2026-06-08
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        iv = int(v)
        if 1000000 <= iv <= 9991231:
            y, m, d = iv // 10000, (iv // 100) % 100, iv % 100
            if y < 1911:
                y += 1911
            try:
                return date(y, m, d).strftime("%Y-%m-%d")
            except ValueError:
                return None
    s = unicodedata.normalize("NFKC", str(v)).strip()
    # 1) ISO อยู่แล้ว (2026-07-15)
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except ValueError:
        pass
    # 2) มีปีเต็ม 4 หลัก: 2026年7月15日 / 2026/7/15 / 2026.7.15
    m2 = _re.search(r"(\d{4})\D{1,2}(\d{1,2})\D{1,2}(\d{1,2})", s)
    if m2:
        try:
            return date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3))).isoformat()
        except ValueError:
            return None
    # 2a) ROC (民國) string: 113/5/13, 115/0507 (ปี 3 หลัก 1xx → +1911) — ต้องอยู่ก่อน 2b กัน 113/5/13 ถูกอ่านเป็น 2013
    roc = _re.fullmatch(r"(1\d{2})\D+(\d{1,2})\D+(\d{1,2})", s) or _re.fullmatch(r"(1\d{2})\D+(\d{2})(\d{2})", s)
    if roc:
        try:
            return date(int(roc.group(1)) + 1911, int(roc.group(2)), int(roc.group(3))).isoformat()
        except ValueError:
            return None
    # 2b) ปีอยู่ท้าย: 7/15/2026, 15/7/2026, 07/15/26 (ค่าใด > 12 = วัน → สลับ; ไม่งั้นถือ M/D)
    m2b = _re.search(r"(\d{1,2})\D{1,2}(\d{1,2})\D{1,2}(\d{2,4})", s)
    if m2b:
        a, b, y = int(m2b.group(1)), int(m2b.group(2)), int(m2b.group(3))
        if y < 100:
            y += 2000
        mo, d = (b, a) if a > 12 else (a, b)
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    # 3) ไม่มีปี — ใช้ year_hint หรือปีปัจจุบัน (เฉพาะเมื่อสตริงไม่มีเลข 4 หลัก)
    if not _re.search(r"\d{4}", s):
        m = _re.search(r"(\d{1,2})\s*[月/\-.]\s*(\d{1,2})\s*日?", s)  # 7月15日 / 7/15
        if m:
            try:
                return date(year_hint or date.today().year, int(m.group(1)), int(m.group(2))).isoformat()
            except ValueError:
                return None
    return None  # อ่านไม่ออก → None (กันส่งค่าขยะเข้า DB)


def _to_pct(v):
    """%欄。自動判斷：>1.5 視為百分比數字 (100, 80, 74.87) → ÷100；<=1.5 視為已是 fraction (0.7887, 0.15, 1)。"""
    n = _to_num(v)
    if n is None:
        return None
    if n > 1.5:
        n = n / 100.0
    return round(n, 6)


def _to_num(v):
    if v is None or v == "":
        return None
    # ตัวเลขอยู่แล้ว → กัน NaN (เซลล์ว่างจาก pandas = float('nan'))
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        return None if math.isnan(f) else f
    # สตริง: ลบ comma / ช่องว่าง / $ / สัญลักษณ์เงิน / % ก่อนแปลง ("1,000,000" "8%" "USD 50,000")
    s = unicodedata.normalize("NFKC", str(v)).strip()
    s = _re.sub(r"[,\s$＄]", "", s)
    s = _re.sub(r"(?i)usd|ntd|twd|jpy|eur", "", s).rstrip("%").strip()
    if not s:
        return None
    try:
        f = float(s)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _looks_dra(coupon, tenor) -> bool:
    """แยก DRA/區間計息 ออกจาก FCN：配息 มี 2 อัตราคั่นด้วย "/" (10%/6.5%) หรือ 期間 เป็นปี (7Y/8Y/10Y)。
    FCN ในไฟล์ = เดือน (7M) + อัตราเดียว → จับแยกได้สะอาด。"""
    if "/" in str(coupon or ""):
        return True
    t = unicodedata.normalize("NFKC", str(tenor or "")).strip().upper()
    return bool(_re.search(r"\d+\s*Y", t))


def _to_bool_mark(v):
    """ช่องเครื่องหมาย (Ｖ/V/✓/是/Y) → True, ว่าง → None, อื่น → False。สำหรับคอลัมน์ BOOLEAN。"""
    if v is None:
        return None
    s = unicodedata.normalize("NFKC", str(v)).strip().upper()
    if s == "":
        return None
    return s in ("V", "✓", "✔", "○", "O", "Y", "YES", "是", "T", "TRUE", "1")


def _norm_name(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s)).replace("＊", "*").strip()
    # ตัดอักขระ PUA (U+E000–U+F8FF) ที่แสดงผลไม่ได้ → ขึ้น "?" (ชื่อหายากในไฟล์ต้นฉบับ เช่น 賴x?)
    return _re.sub(r"[-]", "", s)


def _match_masked(name, candidates):
    """ชื่อย่อมาสก์ (มาสก์ตำแหน่งไหนก็ได้ เช่น 鄭*堃 / 翁* / *美英) → จับคู่ชื่อเต็มความยาวเท่ากัน
    ที่ตัวอักษร "ไม่ถูกมาสก์" ตรงกันทุกตำแหน่ง。คืนเฉพาะเมื่อ match ตัวเดียว (unambiguous) ไม่งั้น None。"""
    if not name:
        return None
    nm = _norm_name(name)
    if "*" not in nm or len(nm) < 2:
        return None
    ln = len(nm)

    def ok(full):
        if not full or len(full) != ln or "*" in full:
            return False
        return all(nm[i] == "*" or nm[i] == full[i] for i in range(ln))

    hits = [c for c in candidates if ok(c)]
    return hits[0] if len(hits) == 1 else None


def _split_code_ticker(code):
    """รหัสที่ ticker หลุดมาติด เช่น EQDS0702773TSLA → (EQDS0702773, TSLA)。ไม่เข้าเงื่อนไข → (code, None)。"""
    if not code:
        return code, None
    mm = _re.match(r"^([A-Z]+\d+)([A-Z]{2,5})$", code)
    if mm:
        return mm.group(1), mm.group(2)
    return code, None


def _clean_ticker(v):
    """全形→半形, 去 $ 前綴, 大寫。"""
    if v is None:
        return None
    t = unicodedata.normalize("NFKC", str(v)).strip().lstrip("$").strip().upper()
    return t or None


def _clean_code(v):
    """全形→半形, 去頭尾空白 (商品代號)。"""
    if v is None:
        return None
    c = unicodedata.normalize("NFKC", str(v)).strip()
    return c or None


def _clean_broker_ticker(v):
    """券商標的 "TSLA UW" / "ORCL UN" → ตัด suffix ตลาด เหลือ ticker。"""
    if v is None:
        return None
    s = unicodedata.normalize("NFKC", str(v)).strip()
    if not s:
        return None
    return _clean_ticker(s.split()[0])


def _is_date_cell(v):
    if isinstance(v, (datetime, date)):
        return True
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return 1000000 <= int(v) <= 9991231
    return False


def _looks_like_code(v):
    """商品代號：含字母、長度>=4、不是「期初價格」。"""
    c = _clean_code(v)
    if not c or _norm(c) == "期初價格":
        return False
    return len(c) >= 4 and any(ch.isalpha() for ch in c)


def _map_headers(rows: list, aliases: dict):
    """หาแถวหัวตาราง + แมพ field -> column index (0-based)。คืน (header_row_idx, {field: col})。
    rows = list ของ tuple (มาจาก iter_rows(values_only=True))。"""
    best_row, best_map = None, {}
    for r in range(min(len(rows), 8)):
        colmap = {}
        for c, val in enumerate(rows[r]):
            h = _norm(val)
            if not h:
                continue
            for field, al in aliases.items():
                if field in colmap:
                    continue
                if any(_norm(a) == h or _norm(a) in h for a in al):
                    colmap[field] = c
                    break
        if len(colmap) > len(best_map):
            best_map, best_row = colmap, r
    return best_row, best_map


def _rows(rows: list, hr: int, colmap: dict):
    out = []
    for r in range(hr + 1, len(rows)):
        row = rows[r]
        rec = {f: (row[c] if c < len(row) else None) for f, c in colmap.items()}
        if all(v is None or str(v).strip() == "" for v in rec.values()):
            continue
        out.append(rec)
    return out


def _sn_block_header(rows: list):
    """ตรวจว่าเป็นชีตแบบ "บล็อก" (DOUU) ไหม。เงื่อนไข: มี 代號+標的1+(執行價/配息/比價)
    และ "ไม่มี" คอลัมน์ 期初價x แยก (ซึ่งจะเป็นเทมเพลตแบนๆ)。คืน (header_row_idx, colmap) หรือ (None, None)。"""
    for r in range(min(len(rows), 6)):
        cm = {}
        has_initial_col = False
        for c, val in enumerate(rows[r]):
            h = _norm(val)
            if not h:
                continue
            if h.startswith("期初") or h.startswith("initial"):
                has_initial_col = True
            for field, al in SN_BLOCK_HDR.items():
                if field in cm:
                    continue
                if any(_norm(a) == h for a in al):
                    cm[field] = c
                    break
        if has_initial_col:
            return None, None  # 有「期初價x」獨立欄 → 是標準範本(扁平)，非區塊
        if "code" in cm and "u1" in cm and ("strike" in cm or "coupon" in cm or "obs" in cm):
            return r, cm
    return None, None


def _parse_sn_block(rows: list, hr: int, cm: dict):
    """อ่านชีตแบบบล็อก → (products, investments)。
    โครงสร้าง: แถวสินค้า (日期+代號) → แถว 期初價格/แถวว่างที่มีราคา → แถวผู้ลงทุน (ชื่อ+ยอด)。"""
    products, investments = [], []
    ucols = [cm[k] for k in ("u1", "u2", "u3", "u4", "u5") if k in cm]
    dcol = cm.get("date", 0)
    ccol = cm.get("code", 1)

    def g(row, key):
        c = cm.get(key)
        return row[c] if (c is not None and c < len(row)) else None

    def has_price(row):
        return any(_to_num(row[c]) is not None for c in ucols if c < len(row))

    def set_initials(prod, row):
        for i, c in enumerate(ucols):
            if c < len(row):
                p = _to_num(row[c])
                if p is not None:
                    prod[f"initial_price_{i + 1}"] = p
        prod["_init"] = True

    cur = None
    for r in range(hr + 1, len(rows)):
        row = rows[r]
        d = row[dcol] if dcol < len(row) else None
        code_cell = row[ccol] if ccol < len(row) else None

        # 期初價格 列 (มี label 期初價格 หรือ ไม่มีชื่อ/รหัส แต่มีราคาในคอลัมน์ standory)
        if _norm(code_cell) == "期初價格" or (
            (d is None or str(d).strip() == "") and _to_num(code_cell) is None and has_price(row)
        ):
            if cur:
                set_initials(cur, row)
            continue

        # แถวสินค้าใหม่: มีวันที่ + รหัสที่ดูเป็นรหัส
        if _is_date_cell(d) and _looks_like_code(code_cell):
            _yr = _year_of(d) or date.today().year
            cur = {
                "product_code": _clean_code(code_cell),
                "category": "SN",
                "product_type": "DRA" if _looks_dra(g(row, "coupon"), g(row, "tenor")) else "FCN",
                "trade_date": _to_date(d, _yr),
                "strike_pct": _to_pct(g(row, "strike")),
                "coupon_pct": _to_pct(g(row, "coupon")),
                "observation_date": _to_date(g(row, "obs"), _yr),
                "exit_date": _to_date(g(row, "exit"), _yr),
                "ko_barrier": _to_pct(g(row, "ko")),
                "ki_barrier": _to_pct(g(row, "ki")),
                "coupon_freq": "monthly",
                "status": "active",
            }
            for i, c in enumerate(ucols):
                cur[f"underlying_{i + 1}"] = _clean_ticker(row[c]) if c < len(row) else None
            # ticker หลุดมาติดรหัส (เช่น EQDS0702773TSLA + 標的1 ว่าง) → แยกออก
            if not cur.get("underlying_1"):
                base, tk = _split_code_ticker(cur["product_code"])
                if tk:
                    cur["product_code"], cur["underlying_1"] = base, tk
            products.append(cur)
            continue

        # แถวผู้ลงทุน: ชื่ออยู่คอลัมน์วันที่ + ยอด(ตัวเลข)อยู่คอลัมน์รหัส
        if cur and d is not None and str(d).strip() != "" and not _is_date_cell(d) and _to_num(code_cell) is not None:
            investments.append({
                "customer": unicodedata.normalize("NFKC", str(d)).strip(),
                "product_code": cur["product_code"],
                "amount_usd": _to_num(code_cell) or 0,
                "currency": "USD",
            })
            # แถวผู้ลงทุนแรกมักพ่วงราคาเริ่มต้นมาด้วย
            if not cur.get("_init") and has_price(row):
                set_initials(cur, row)
            continue

    for p in products:
        p.pop("_init", None)
    return products, investments


def _broker_header(rows: list):
    """ตรวจชีต "券商庫存匯出" (扁平)。ต้องมี code + customer + amount。คืน (hr, colmap) หรือ (None, None)。"""
    hr, cm = _map_headers(rows, BROKER_HDR)
    if cm and all(k in cm for k in ("code", "customer", "amount")):
        return hr, cm
    return None, None


def _parse_broker(rows: list, hr: int, cm: dict):
    """อ่านชีตแบบ券商庫存 (แถวละ 1 持倉) → (products, customers, investments)。
    商品 dedup ตาม code; 客戶 dedup ตามชื่อ。配息起算=發行日, 配息頻率=月配。"""
    products, customers, investments = {}, {}, []
    for rec in _rows(rows, hr, cm):
        code = _clean_code(rec.get("code"))
        if not code or _norm(rec.get("code")) in ("psc商品代碼", "商品代碼"):
            continue
        cust = _norm_name(rec.get("customer"))
        amt = _to_num(rec.get("amount"))
        cur = (str(rec.get("currency")).strip() if rec.get("currency") else "") or "USD"

        if code not in products:
            pt = (str(rec.get("ptype")).strip().upper() if rec.get("ptype") else "FCN")
            p = {
                "product_code": code,
                "category": "SN",
                "product_type": "DRA" if pt == "DRA" else "FCN",
                "trade_date": _to_date(rec.get("trade_date")),
                "observation_date": _to_date(rec.get("obs_date")),   # 發行日
                "exit_date": _to_date(rec.get("exit_date")),
                "coupon_pct": _to_pct(rec.get("coupon")),
                "ko_barrier": _to_pct(rec.get("ko")),
                "ki_barrier": _to_pct(rec.get("ki")),
                "strike_pct": _to_pct(rec.get("strike")),
                "coupon_freq": "monthly",
                "status": "active",
            }
            for i in range(1, 6):
                tk = _clean_broker_ticker(rec.get(f"u{i}"))
                p[f"underlying_{i}"] = tk
                p[f"initial_price_{i}"] = _to_num(rec.get(f"ip{i}")) if tk else None
            products[code] = p

        if cust and cust not in customers:
            customers[cust] = {"name": cust}
        if cust and amt:
            investments.append({"customer": cust, "product_code": code,
                                "amount_usd": amt, "currency": cur})
    return list(products.values()), list(customers.values()), investments


def _materialize(ws, max_empty: int = 50):
    """อ่านแถวจาก read-only sheet แบบมีขอบเขต (กัน max_row บานเป็นล้านแถวจาก format ค้าง)。"""
    out, empty = [], 0
    for row in ws.iter_rows(values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            empty += 1
            if empty > max_empty:
                break
        else:
            empty = 0
        out.append(row)
    while out and all(v is None or str(v).strip() == "" for v in out[-1]):
        out.pop()
    return out


def parse_workbook(data: bytes) -> dict:
    # read_only=True → อ่านแบบ stream ไม่โหลด style ทั้งไฟล์ (ไฟล์บวม format จะเร็วขึ้นมาก, กัน OOM/timeout บน Render)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    customers, products, investments, warnings = [], [], [], []

    sheets = {sn: _materialize(wb[sn]) for sn in wb.sheetnames}
    wb.close()

    # 0) ชีตแบบ "券商庫存匯出" (扁平 แถวละ 1 持倉 รวม สินค้า+ลูกค้า+ยอด) → จบในตัว
    for sn, rows in sheets.items():
        hr, cm = _broker_header(rows)
        if cm is not None:
            p, c, iv = _parse_broker(rows, hr, cm)
            products += p
            customers += c
            investments += iv
    if products or customers or investments:
        return {"customers": customers, "products": products, "investments": investments, "warnings": warnings}

    # 1) ชีตแบบ "บล็อก" (ฟอร์แมต DOUU ปัจจุบัน) → สินค้า + การลงทุน
    block_handled = False
    for sn, rows in sheets.items():
        hr, cm = _sn_block_header(rows)
        if cm is not None:
            p, iv = _parse_sn_block(rows, hr, cm)
            products += p
            investments += iv
            block_handled = True

    # 2) จับชีตตามชื่อ (เทมเพลตมาตรฐาน 客戶/商品/投資 + 開戶明細)
    kind_sheets = {"cust": None, "prod": None, "inv": None}
    for sn in sheets:
        n = _norm(sn)
        for hint, kind in SHEET_HINT.items():
            if _norm(hint) in n and kind_sheets[kind] is None:
                kind_sheets[kind] = sn

    # 客戶 (ทุกฟอร์แมต — 開戶明細 หรือ 客戶 sheet)
    if kind_sheets["cust"]:
        rows = sheets[kind_sheets["cust"]]
        hr, cm = _map_headers(rows, CUST)
        if "name" in cm:
            for rec in _rows(rows, hr, cm):
                nm = rec.get("name")
                if not nm or _norm(nm) in ("姓名", "戶名") or _to_num(nm) is not None:
                    continue
                customers.append({
                    "name": unicodedata.normalize("NFKC", str(nm)).strip(),
                    "usd_amount": _to_num(rec.get("usd_amount")),
                    "unified_account": _to_bool_mark(rec.get("unified_account")),
                    "pi_signed": _to_bool_mark(rec.get("pi_signed")),
                    "ordered": _to_bool_mark(rec.get("ordered")),
                    "ctbc_position": _to_num(rec.get("ctbc_position")),
                    "fund_amount": _to_num(rec.get("fund_amount")),
                    "notes": (str(rec.get("notes")).strip() if rec.get("notes") else None),
                })

    # 商品 / 投資 แบบเทมเพลตแบนๆ — เฉพาะตอนไม่เจอบล็อก
    if not block_handled and kind_sheets["prod"]:
        rows = sheets[kind_sheets["prod"]]
        hr, cm = _map_headers(rows, PROD)
        if "product_code" in cm:
            for rec in _rows(rows, hr, cm):
                code = rec.get("product_code")
                if not code or _norm(code) in ("代號", "商品代號", "期初價格"):
                    continue
                fr = rec.get("coupon_freq")
                st = rec.get("status")
                products.append({
                    "product_code": _clean_code(code),
                    "category": (str(rec.get("category")).strip() if rec.get("category") else "SN"),
                    "underlying_1": _clean_ticker(rec.get("underlying_1")),
                    "underlying_2": _clean_ticker(rec.get("underlying_2")),
                    "underlying_3": _clean_ticker(rec.get("underlying_3")),
                    "initial_price_1": _to_num(rec.get("initial_price_1")),
                    "initial_price_2": _to_num(rec.get("initial_price_2")),
                    "initial_price_3": _to_num(rec.get("initial_price_3")),
                    "ko_barrier": _to_pct(rec.get("ko_barrier")),
                    "ki_barrier": _to_pct(rec.get("ki_barrier")),
                    "strike_pct": _to_pct(rec.get("strike_pct")),
                    "coupon_pct": _to_pct(rec.get("coupon_pct")),
                    "coupon_freq": FREQ_MAP.get(str(fr).strip(), "monthly") if fr else "monthly",
                    "trade_date": _to_date(rec.get("trade_date")),
                    "observation_date": _to_date(rec.get("observation_date")),
                    "exit_date": _to_date(rec.get("exit_date")),
                    "status": STATUS_MAP.get(str(st).strip(), "active") if st else "active",
                })

    if not block_handled and kind_sheets["inv"]:
        rows = sheets[kind_sheets["inv"]]
        hr, cm = _map_headers(rows, INV)
        if "customer" in cm and "product_code" in cm:
            for rec in _rows(rows, hr, cm):
                cu = rec.get("customer"); pc = rec.get("product_code")
                if not cu or not pc:
                    continue
                investments.append({
                    "customer": unicodedata.normalize("NFKC", str(cu)).strip(),
                    "product_code": _clean_code(pc),
                    "amount_usd": _to_num(rec.get("amount_usd")) or 0,
                    "currency": (str(rec.get("currency")).strip() if rec.get("currency") else "USD"),
                })

    if not products and not customers and not investments:
        warnings.append("找不到可匯入的資料。請使用標準範本，或確認檔案含 開戶/SN 分頁。")

    return {"customers": customers, "products": products, "investments": investments, "warnings": warnings}


_DIFF_PFIELDS = ["underlying_1", "underlying_2", "underlying_3", "underlying_4", "underlying_5",
                 "initial_price_1", "initial_price_2", "initial_price_3", "initial_price_4", "initial_price_5",
                 "strike_pct", "coupon_pct", "ko_barrier", "ki_barrier",
                 "trade_date", "observation_date", "exit_date", "status"]
_DIFF_CFIELDS = ["usd_amount", "ctbc_position", "fund_amount", "unified_account", "pi_signed", "ordered"]


def _changed(old, new):
    """ค่าใหม่ "จะทับ" ค่าเก่าจริงไหม。new None → ไม่ทับ (commit ข้าม None) → ไม่ใช่ change。"""
    if new is None or new == "":
        return False
    on, nn = _to_num(old), _to_num(new)
    if nn is not None and on is not None:
        return abs(on - nn) > 1e-6
    so = "" if old is None else str(old).strip()
    return so != str(new).strip()


def _diff_existing(r: Repo, parsed: dict) -> dict:
    """เทียบกับข้อมูลเดิมในระบบ → คืนช่องที่ "จะเปลี่ยน" (เก่า→ใหม่) ให้ผู้ใช้ตรวจก่อนยืนยัน。"""
    out = {"products": [], "customers": []}
    try:
        ex_p = {p.get("product_code"): p for p in r.list("structured_notes")}
        for p in parsed.get("products", []):
            ex = ex_p.get(p.get("product_code"))
            if not ex:
                continue
            diffs = [{"field": f, "old": ex.get(f), "new": p.get(f)}
                     for f in _DIFF_PFIELDS if _changed(ex.get(f), p.get(f))]
            if diffs:
                out["products"].append({"key": p.get("product_code"), "diffs": diffs})
        ex_c = {c.get("name"): c for c in r.list("customers")}
        for c in parsed.get("customers", []):
            ex = ex_c.get(c.get("name"))
            if not ex:
                continue
            diffs = [{"field": f, "old": ex.get(f), "new": c.get(f)}
                     for f in _DIFF_CFIELDS if _changed(ex.get(f), c.get(f))]
            if diffs:
                out["customers"].append({"key": c.get("name"), "diffs": diffs})
    except Exception:
        pass  # diff ล้มเหลวไม่ควรบล็อก preview
    return out


@router.post("/preview")
async def preview(file: UploadFile = File(...), r: Repo = Depends(repo)):
    try:
        data = await file.read()
        parsed = parse_workbook(data)
        parsed["changes"] = _diff_existing(r, parsed)
        return parsed
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"讀取檔案失敗: {e}")


_PFIELDS = ("product_code", "category",
            "underlying_1", "underlying_2", "underlying_3", "underlying_4", "underlying_5",
            "initial_price_1", "initial_price_2", "initial_price_3", "initial_price_4", "initial_price_5",
            "ko_barrier", "ki_barrier",
            "strike_pct", "coupon_pct", "coupon_freq", "trade_date", "observation_date", "exit_date", "status")


@router.post("/commit")
def commit(body: dict, r: Repo = Depends(repo)):
    """寫入登入者自己的 tenant，去重 (依姓名/代號)，連結投資。
    一次載入既有資料 + 批次寫入 (少 round-trip → 不會在 Render 逾時/斷線)。"""
    try:
        return _do_commit(body, r)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"匯入失敗: {getattr(e, 'message', None) or e}")


def _do_commit(body: dict, r: Repo) -> dict:
    from .products import _safe_write  # reuse column-resilient single writer
    custs = body.get("customers") or []
    prods = body.get("products") or []
    invs = body.get("investments") or []
    res = {"customers_created": 0, "customers_updated": 0,
           "products_created": 0, "products_updated": 0,
           "investments_created": 0, "investments_skipped": 0, "warnings": []}

    # 一次載入既有資料，建索引 (取代逐筆 find)
    name2id = {c["name"]: c["id"] for c in r.list("customers", select="id,name") if c.get("name")}
    code2id = {s["product_code"]: s["id"] for s in r.list("structured_notes", select="id,product_code") if s.get("product_code")}
    inv_set = {(i.get("customer_id"), i.get("sn_id")) for i in r.list("investments", select="customer_id,sn_id")}

    # 客戶：既有 → update；新 → 批次 insert
    new_cust = []
    for cu in custs:
        nm = (cu.get("name") or "").strip()
        if not nm:
            continue
        payload = {k: cu.get(k) for k in ("name", "usd_amount", "unified_account", "pi_signed",
                                          "ordered", "ctbc_position", "fund_amount", "notes") if cu.get(k) is not None}
        payload["name"] = nm
        if nm in name2id:
            _safe_write(lambda b, _id=name2id[nm]: r.update("customers", _id, b), payload)
            res["customers_updated"] += 1
        else:
            new_cust.append(payload)
    for row in _safe_bulk(r, "customers", new_cust):
        if row.get("name"):
            name2id[row["name"]] = row["id"]
    res["customers_created"] += len(new_cust)

    # 商品：逐筆 (保留 column-resilient)，但不再 find
    for p in prods:
        code = (p.get("product_code") or "").strip()
        if not code:
            continue
        payload = {k: p.get(k) for k in _PFIELDS if p.get(k) is not None}
        if code in code2id:
            _safe_write(lambda b, _id=code2id[code]: r.update("structured_notes", _id, b), payload)
            res["products_updated"] += 1
        else:
            pid = _safe_write(lambda b: r.create("structured_notes", b), payload)["id"]
            code2id[code] = pid
            res["products_created"] += 1

    # 投資：解析投資人 → exact → จับคู่ชื่อย่อ X*Y กับชื่อเต็ม → ที่เหลือค่อยสร้างใหม่
    missing, seen = [], set()
    for iv in invs:
        nm = (iv.get("customer") or "").strip()
        code = (iv.get("product_code") or "").strip()
        if not (code in code2id and nm) or nm in name2id or nm in seen:
            continue
        mt = _match_masked(nm, list(name2id.keys()))
        if mt:
            name2id[nm] = name2id[mt]   # ชื่อย่อ → id ลูกค้าจริง (ไม่สร้างซ้ำ)
            res["warnings"].append(f"投資人「{nm}」自動對應到既有客戶「{mt}」")
            continue
        missing.append(nm); seen.add(nm)
    if missing:
        for row in _safe_bulk(r, "customers", [{"name": nm} for nm in missing]):
            if row.get("name"):
                name2id[row["name"]] = row["id"]
        res["customers_created"] += len(missing)
        for nm in missing:
            res["warnings"].append(f"自動建立投資人「{nm}」— 可能是既有客戶的暱稱，請至客戶頁確認/合併")

    new_invs = []
    for iv in invs:
        nm = (iv.get("customer") or "").strip()
        code = (iv.get("product_code") or "").strip()
        pid = code2id.get(code)
        cid = name2id.get(nm)
        if not pid:
            res["investments_skipped"] += 1
            res["warnings"].append(f"連結失敗(找不到商品): {iv.get('customer')} / {code}")
            continue
        if not cid or (cid, pid) in inv_set:
            res["investments_skipped"] += 1
            continue
        inv_set.add((cid, pid))
        new_invs.append({"customer_id": cid, "sn_id": pid,
                         "amount_usd": iv.get("amount_usd") or 0})
    _safe_bulk(r, "investments", new_invs)
    res["investments_created"] += len(new_invs)

    return res


@router.post("/cleanup")
def cleanup_dupes(r: Repo = Depends(repo)):
    """整理重複（deterministic）：暱稱 X*Y 合併到同名全名客戶；代號夾帶 ticker 合併到正確代號。
    先把投資轉到正確的人/商品（去重）再刪掉重複者。"""
    try:
        return _do_cleanup(r)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"整理失敗: {getattr(e, 'message', None) or e}")


def _merge_customer(r, from_id, to_id, invs, pairs):
    """ย้ายการลงทุนของ from_id → to_id (ตัดที่ซ้ำ) แล้วลบลูกค้า from_id。"""
    for i in [x for x in invs if x.get("customer_id") == from_id]:
        if (to_id, i.get("sn_id")) in pairs:
            r.delete("investments", i["id"])
        else:
            r.update("investments", i["id"], {"customer_id": to_id})
            pairs.add((to_id, i.get("sn_id")))
            i["customer_id"] = to_id
    r.delete("customers", from_id)


def _do_cleanup(r: Repo) -> dict:
    res = {"customers_merged": 0, "products_merged": 0, "details": []}

    invs = r.list("investments", select="id,customer_id,sn_id")
    pairs = {(i.get("customer_id"), i.get("sn_id")) for i in invs}

    # 1a) 合併「完全同名」客戶 (เช่น 翁* ซ้ำ 2 แถว)
    custs = r.list("customers", select="id,name")
    by_name = {}
    for c in custs:
        key = _norm_name(c.get("name"))
        if not key:
            continue
        if key in by_name:
            _merge_customer(r, c["id"], by_name[key], invs, pairs)
            res["customers_merged"] += 1
            res["details"].append(f"{c.get('name')} (重複) → 合併")
        else:
            by_name[key] = c["id"]

    # 1b) 合併暱稱客戶 (มาสก์ → 全名)
    custs = r.list("customers", select="id,name")
    full_names = {c["name"]: c["id"] for c in custs if c.get("name") and "*" not in _norm_name(c["name"])}
    for c in custs:
        nm = c.get("name") or ""
        mt = _match_masked(nm, list(full_names.keys()))
        if not mt:
            continue
        fid = full_names[mt]
        if fid == c["id"]:
            continue
        _merge_customer(r, c["id"], fid, invs, pairs)
        res["customers_merged"] += 1
        res["details"].append(f"{nm} → {mt}")

    # 2) 合併代號夾帶 ticker (EQDS..TSLA → EQDS..)
    prods = r.list("structured_notes", select="id,product_code")
    code2id = {p["product_code"]: p["id"] for p in prods if p.get("product_code")}
    invs = r.list("investments", select="id,customer_id,sn_id")
    pairs = {(i.get("customer_id"), i.get("sn_id")) for i in invs}
    for p in prods:
        code = p.get("product_code") or ""
        base, tk = _split_code_ticker(code)
        if not tk or base == code or base not in code2id:
            continue
        clean_id = code2id[base]
        if clean_id == p["id"]:
            continue
        for i in [x for x in invs if x.get("sn_id") == p["id"]]:
            if (i.get("customer_id"), clean_id) in pairs:
                r.delete("investments", i["id"])
            else:
                r.update("investments", i["id"], {"sn_id": clean_id})
                pairs.add((i.get("customer_id"), clean_id))
        r.delete("structured_notes", p["id"])
        res["products_merged"] += 1
        res["details"].append(f"{code} → {base}")

    return res
