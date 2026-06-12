"""資料匯出 — 把登入者 tenant 的 客戶/商品/投資 匯出成 Excel (與匯入範本同格式，可再匯入)。"""
import io
from datetime import date
from fastapi import APIRouter, Depends, Response
from ..deps import repo
from ..db import Repo

router = APIRouter(prefix="/api/export", tags=["export"])

FREQ_REV = {"monthly": "月配", "quarterly": "季配", "semiannual": "半年配", "annual": "年配", "maturity": "到期一次"}
STATUS_REV = {"active": "進行中", "exited": "已出場", "inactive": "暫停"}
GREEN = "1F4D3A"; GREEN_DK = "153A2B"; ZEBRA = "F4F9F6"; BORDER_C = "D4E5DA"


def _pct(v):
    return round(float(v) * 100, 4) if isinstance(v, (int, float)) else None


def _d(v):
    return str(v)[:10] if v else None


@router.get("/excel")
def export_excel(r: Repo = Depends(repo)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    custs = r.list("customers", order="name")
    prods = r.list("structured_notes", order="product_code")
    invs = r.find("investments", select="amount_usd,currency,customers(name),structured_notes(product_code)")

    thin = Side(style="thin", color=BORDER_C)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor=GREEN_DK)
    banner_fill = PatternFill("solid", fgColor=GREEN)
    head_font = Font(bold=True, color="FFFFFF")

    wb = Workbook(); wb.remove(wb.active)

    def sheet(title, headers, rows, widths, nfmts=None):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        last = get_column_letter(len(headers))
        ws.merge_cells(f"A1:{last}1")
        ws["A1"] = f"  Justinvestment · {title}"
        ws["A1"].fill = banner_fill; ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws.row_dimensions[1].height = 28
        for i, h in enumerate(headers, 1):
            c = ws.cell(2, i, h); c.fill = head_fill; c.font = head_font; c.border = border
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
        ws.freeze_panes = "A3"
        for ri, row in enumerate(rows, 3):
            fill = PatternFill("solid", fgColor="FFFFFF" if ri % 2 else ZEBRA)
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val); c.fill = fill; c.border = border
                if nfmts and nfmts[ci - 1]:
                    c.number_format = nfmts[ci - 1]
        return ws

    # 客戶
    sheet("客戶", ["姓名", "美元額度", "幣別", "統一帳號", "備註"],
          [[c.get("name"), c.get("usd_amount"), c.get("currency") or "USD",
            c.get("unified_account"), c.get("notes")] for c in custs],
          [18, 14, 9, 16, 32], [None, "#,##0", None, None, None])

    # 商品
    sheet("商品",
          ["商品代號", "類別", "標的1", "期初價1", "標的2", "期初價2", "標的3", "期初價3",
           "KO障壁(%)", "KI障壁(%)", "履約價(%)", "票息(%年化)", "配息頻率", "成交日", "比價日", "到期日", "狀態"],
          [[p.get("product_code"), p.get("category") or "SN",
            p.get("underlying_1"), p.get("initial_price_1"),
            p.get("underlying_2"), p.get("initial_price_2"),
            p.get("underlying_3"), p.get("initial_price_3"),
            _pct(p.get("ko_barrier")), _pct(p.get("ki_barrier")), _pct(p.get("strike_pct")),
            _pct(p.get("coupon_pct")), FREQ_REV.get(p.get("coupon_freq") or "monthly", "月配"),
            _d(p.get("trade_date")), _d(p.get("observation_date")), _d(p.get("exit_date")),
            STATUS_REV.get(p.get("status") or "active", "進行中")] for p in prods],
          [16, 10, 9, 10, 9, 10, 9, 10, 11, 11, 11, 12, 11, 12, 12, 12, 10],
          [None, None, None, "#,##0.00", None, "#,##0.00", None, "#,##0.00",
           "0", "0", "0", "0", None, None, None, None, None])

    # 投資
    sheet("投資", ["客戶姓名", "商品代號", "投資金額", "幣別"],
          [[(iv.get("customers") or {}).get("name"),
            (iv.get("structured_notes") or {}).get("product_code"),
            iv.get("amount_usd"), iv.get("currency") or "USD"] for iv in invs],
          [18, 16, 14, 9], [None, None, "#,##0", None])

    buf = io.BytesIO(); wb.save(buf)
    name = f"investment_export_{date.today():%Y%m%d}.xlsx"
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{name}"'})
